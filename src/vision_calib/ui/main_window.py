"""
主應用程式視窗

提供 vision-calib 的主要圖形介面，整合所有標定功能。
採用 Google Material Design 3 設計語言。
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Qt, QSize, Signal, Slot, QThread, QMutex, QMutexLocker
from PySide6.QtGui import QAction, QIcon, QFont, QImage
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QSizePolicy,
    QSpacerItem,
    QSplitter,
    QStatusBar,
    QTabWidget,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

from vision_calib import __version__
from vision_calib.core.types import CalibrationResult
from vision_calib.io import CalibrationFile
from vision_calib.ui.styles.theme import Theme, ThemeManager
from vision_calib.utils.logging import get_logger, setup_logging

logger = get_logger("ui.main_window")


class ImageViewer(QLabel):
    """
    支持縮放、拖曳、標記點的圖像查看器

    功能：
    - 左鍵點擊：新增標記點 / 選中現有點
    - 右鍵拖曳：平移圖像
    - 滾輪：縮放
    - 雙擊：重置視圖
    """

    point_added = Signal(float, float)  # 新增點 (image_x, image_y)
    point_selected = Signal(int)  # 選中點 (index)
    mouse_moved = Signal(float, float)  # 滑鼠移動 (image_x, image_y)
    mouse_left = Signal()  # 滑鼠離開

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumSize(640, 480)
        self.setMouseTracking(True)
        self.setStyleSheet("""
            ImageViewer {
                background-color: #1a1a1a;
                border: 1px solid #404040;
                border-radius: 4px;
            }
        """)
        self.setText("請選擇相機並點擊「開始串流」")

        # 圖像數據
        self._pixmap = None  # 原始圖像
        self._image_width = 0
        self._image_height = 0

        # 視圖變換
        self._scale = 1.0
        self._offset_x = 0.0
        self._offset_y = 0.0
        self._min_scale = 0.1
        self._max_scale = 10.0

        # 拖曳狀態
        self._dragging = False
        self._drag_start = None
        self._drag_offset_start = None

        # 標記點
        self._points = []  # [{name, x, y}, ...]
        self._point_radius = 15  # 點擊檢測半徑

        # 模式
        self._interactive = False  # 是否允許交互（拍照後才允許）

    def set_image(self, pixmap, interactive: bool = False):
        """設置要顯示的圖像"""
        from PySide6.QtGui import QPixmap
        self._pixmap = pixmap
        if pixmap:
            self._image_width = pixmap.width()
            self._image_height = pixmap.height()
        self._interactive = interactive
        self._update_display()

    def set_points(self, points: list):
        """設置標記點列表"""
        self._points = points
        self._update_display()

    def reset_view(self):
        """重置視圖（適應窗口）"""
        self._scale = 1.0
        self._offset_x = 0.0
        self._offset_y = 0.0
        self._update_display()

    def _update_display(self):
        """更新顯示"""
        from PySide6.QtGui import QPixmap, QPainter, QPen, QColor, QFont

        if self._pixmap is None:
            return

        # 計算適應窗口的基礎縮放
        label_w, label_h = self.width(), self.height()
        if label_w <= 0 or label_h <= 0:
            return

        base_scale = min(label_w / self._image_width, label_h / self._image_height)
        total_scale = base_scale * self._scale

        # 計算縮放後的尺寸
        scaled_w = int(self._image_width * total_scale)
        scaled_h = int(self._image_height * total_scale)

        # 創建顯示用的 pixmap
        display = QPixmap(label_w, label_h)
        display.fill(QColor("#1a1a1a"))

        painter = QPainter(display)
        painter.setRenderHint(QPainter.SmoothPixmapTransform)

        # 計算繪製位置（居中 + 偏移）
        draw_x = (label_w - scaled_w) / 2 + self._offset_x
        draw_y = (label_h - scaled_h) / 2 + self._offset_y

        # 繪製圖像
        painter.drawPixmap(
            int(draw_x), int(draw_y), scaled_w, scaled_h,
            self._pixmap
        )

        # 繪製標記點
        if self._interactive and self._points:
            for i, point in enumerate(self._points):
                # 轉換圖像座標到顯示座標
                px = draw_x + point['pixel_x'] * total_scale
                py = draw_y + point['pixel_y'] * total_scale

                # 繪製十字線
                pen = QPen(QColor(255, 50, 50), 2)
                painter.setPen(pen)
                cross_size = 12
                painter.drawLine(int(px - cross_size), int(py), int(px + cross_size), int(py))
                painter.drawLine(int(px), int(py - cross_size), int(px), int(py + cross_size))

                # 繪製圓圈
                painter.drawEllipse(int(px - 8), int(py - 8), 16, 16)

                # 繪製標籤
                painter.setFont(QFont("Arial", 11, QFont.Bold))
                painter.setPen(QPen(QColor(255, 255, 0)))
                painter.drawText(int(px + 12), int(py - 5), point['name'])

        painter.end()
        super().setPixmap(display)

    def _display_to_image_coords(self, display_x: float, display_y: float):
        """將顯示座標轉換為圖像座標"""
        if self._pixmap is None or self._image_width == 0:
            return None, None

        label_w, label_h = self.width(), self.height()
        base_scale = min(label_w / self._image_width, label_h / self._image_height)
        total_scale = base_scale * self._scale

        scaled_w = self._image_width * total_scale
        scaled_h = self._image_height * total_scale

        draw_x = (label_w - scaled_w) / 2 + self._offset_x
        draw_y = (label_h - scaled_h) / 2 + self._offset_y

        img_x = (display_x - draw_x) / total_scale
        img_y = (display_y - draw_y) / total_scale

        if 0 <= img_x < self._image_width and 0 <= img_y < self._image_height:
            return img_x, img_y
        return None, None

    def _find_point_at(self, img_x: float, img_y: float):
        """查找指定位置附近的點，返回索引或 -1"""
        import math
        for i, point in enumerate(self._points):
            dist = math.sqrt((img_x - point['pixel_x'])**2 + (img_y - point['pixel_y'])**2)
            if dist < self._point_radius:
                return i
        return -1

    def mousePressEvent(self, event):
        """滑鼠按下"""
        if event.button() == Qt.RightButton:
            # 右鍵開始拖曳
            self._dragging = True
            self._drag_start = event.pos()
            self._drag_offset_start = (self._offset_x, self._offset_y)
            self.setCursor(Qt.ClosedHandCursor)
        elif event.button() == Qt.LeftButton and self._interactive:
            # 左鍵點擊標記點
            img_x, img_y = self._display_to_image_coords(event.pos().x(), event.pos().y())
            if img_x is not None:
                # 檢查是否點擊了現有點
                point_idx = self._find_point_at(img_x, img_y)
                if point_idx >= 0:
                    self.point_selected.emit(point_idx)
                else:
                    self.point_added.emit(img_x, img_y)
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        """滑鼠釋放"""
        if event.button() == Qt.RightButton:
            self._dragging = False
            self.setCursor(Qt.ArrowCursor)
        super().mouseReleaseEvent(event)

    def mouseMoveEvent(self, event):
        """滑鼠移動"""
        if self._dragging and self._drag_start:
            # 拖曳中
            delta = event.pos() - self._drag_start
            self._offset_x = self._drag_offset_start[0] + delta.x()
            self._offset_y = self._drag_offset_start[1] + delta.y()
            self._update_display()
        elif self._interactive:
            # 顯示座標
            img_x, img_y = self._display_to_image_coords(event.pos().x(), event.pos().y())
            if img_x is not None:
                self.mouse_moved.emit(img_x, img_y)
            else:
                self.mouse_left.emit()
        super().mouseMoveEvent(event)

    def leaveEvent(self, event):
        """滑鼠離開"""
        self.mouse_left.emit()
        super().leaveEvent(event)

    def wheelEvent(self, event):
        """滾輪縮放（以滑鼠位置為中心）"""
        if self._pixmap is None:
            return

        # 獲取滾輪方向
        delta = event.angleDelta().y()
        factor = 1.15 if delta > 0 else 1 / 1.15

        # 計算新縮放
        old_scale = self._scale
        new_scale = self._scale * factor
        new_scale = max(self._min_scale, min(self._max_scale, new_scale))

        if new_scale != old_scale:
            # 獲取滑鼠位置
            mouse_x = event.position().x()
            mouse_y = event.position().y()

            # 計算當前圖像中心位置
            label_w, label_h = self.width(), self.height()
            base_scale = min(label_w / self._image_width, label_h / self._image_height)

            # 縮放前的圖像繪製位置
            old_total_scale = base_scale * old_scale
            old_scaled_w = self._image_width * old_total_scale
            old_scaled_h = self._image_height * old_total_scale
            old_draw_x = (label_w - old_scaled_w) / 2 + self._offset_x
            old_draw_y = (label_h - old_scaled_h) / 2 + self._offset_y

            # 滑鼠位置對應的圖像座標
            img_x = (mouse_x - old_draw_x) / old_total_scale
            img_y = (mouse_y - old_draw_y) / old_total_scale

            # 應用新縮放
            self._scale = new_scale
            new_total_scale = base_scale * new_scale
            new_scaled_w = self._image_width * new_total_scale
            new_scaled_h = self._image_height * new_total_scale

            # 計算新的繪製位置（使同一圖像點保持在滑鼠位置）
            new_draw_x = mouse_x - img_x * new_total_scale
            new_draw_y = mouse_y - img_y * new_total_scale

            # 更新偏移
            self._offset_x = new_draw_x - (label_w - new_scaled_w) / 2
            self._offset_y = new_draw_y - (label_h - new_scaled_h) / 2

            self._update_display()

        event.accept()

    def mouseDoubleClickEvent(self, event):
        """雙擊重置視圖"""
        self.reset_view()
        super().mouseDoubleClickEvent(event)

    def resizeEvent(self, event):
        """視窗大小改變時重新繪製"""
        super().resizeEvent(event)
        self._update_display()


class WebCamWorker(QThread):
    """
    背景執行緒處理 WebCAM 影像擷取

    架構：
    - Signal 只傳輕量 QImage（預覽用）
    - 原始幀通過 get_frame() 獲取（線程安全）
    - 主執行緒零 CPU 負擔
    """

    preview_ready = Signal(QImage)  # 只傳預覽圖，不傳原始幀
    resolution_ready = Signal(int, int)
    error = Signal(str)

    def __init__(self, cam_index: int, parent=None):
        super().__init__(parent)
        self._cam_index = cam_index
        self._running = False
        self._cap = None
        self._frame_lock = QMutex()
        self._latest_frame = None  # BGR 原始幀
        self._frame_width = 0
        self._frame_height = 0

    def get_frame(self):
        """線程安全地獲取最新原始幀（用於拍照）"""
        with QMutexLocker(self._frame_lock):
            if self._latest_frame is not None:
                return self._latest_frame.copy(), self._frame_width, self._frame_height
            return None, 0, 0

    def run(self):
        """執行緒主迴圈"""
        import cv2
        import time
        import numpy as np

        # 開啟相機
        self._cap = cv2.VideoCapture(self._cam_index, cv2.CAP_DSHOW)
        if not self._cap.isOpened():
            self.error.emit(f"無法開啟相機 {self._cam_index}")
            return

        # 設置最高解析度
        self._cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self._cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self._cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        # 獲取實際解析度
        self._frame_width = int(self._cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self._frame_height = int(self._cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.resolution_ready.emit(self._frame_width, self._frame_height)

        # 幀率控制（15 FPS 預覽足夠，降低 CPU 負擔）
        frame_interval = 1.0 / 15
        last_emit_time = 0

        # 預先計算縮放參數
        preview_w = 640
        scale = preview_w / self._frame_width
        preview_h = int(self._frame_height * scale)

        self._running = True

        while self._running:
            ret, frame = self._cap.read()
            if not ret:
                time.sleep(0.01)
                continue

            # 保存原始幀（線程安全）
            with QMutexLocker(self._frame_lock):
                self._latest_frame = frame

            # 幀率控制
            current_time = time.time()
            if current_time - last_emit_time < frame_interval:
                time.sleep(0.001)  # 避免忙等待
                continue
            last_emit_time = current_time

            # 快速縮放（使用 INTER_NEAREST 最快）
            small = cv2.resize(frame, (preview_w, preview_h), interpolation=cv2.INTER_NEAREST)

            # BGR 轉 RGB（使用 numpy 切片比 cvtColor 快）
            rgb = small[:, :, ::-1].copy()

            # 創建 QImage（使用 copy() 確保數據獨立）
            q_img = QImage(
                rgb.data,
                preview_w, preview_h,
                3 * preview_w,
                QImage.Format_RGB888
            ).copy()

            self.preview_ready.emit(q_img)

        # 清理
        if self._cap is not None:
            self._cap.release()
            self._cap = None

    def stop(self):
        """停止執行緒"""
        self._running = False
        self.wait(2000)


class MainWindow(QMainWindow):
    """主應用程式視窗"""

    def __init__(self):
        super().__init__()

        # 主題管理器
        self.theme_manager = ThemeManager()

        self.setWindowTitle(f"TSIC/CR-ICS01 相機與手臂整合控制教學軟體 v{__version__}")
        self.setMinimumSize(800, 500)
        self.resize(1280, 800)

        # 當前標定結果
        self._result: Optional[CalibrationResult] = None

        # 背景工作執行緒
        self._corner_worker = None
        self._calib_worker = None

        # 角點偵測結果快取 {image_path: corners}
        self._corner_cache: dict = {}

        # 外參標定結果
        self._extrinsic_result = None

        # 座標轉換器
        self._transformer = None

        # 點位數據 [id, image_x, image_y, world_x, world_y]
        self._point_data: list = []

        # 生成的世界座標
        self._generated_world_coords: list = []

        # WebCAM 標記點 [{name, pixel_x, pixel_y, world_x, world_y}, ...]
        self._marked_points: list = []

        # 設置 UI
        self._setup_ui()
        self._setup_menu()
        self._setup_toolbar()
        self._setup_statusbar()

        # 套用主題
        self.theme_manager.apply_current_theme()

        logger.info("主視窗初始化完成")

    def _setup_ui(self):
        """設置主要 UI 佈局"""
        # 中央組件
        central = QWidget()
        self.setCentralWidget(central)

        # 主佈局
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 分割器
        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(1)

        # 左側面板 - 控制區
        left_panel = self._create_control_panel()
        splitter.addWidget(left_panel)

        # 右側面板 - 標籤頁
        right_panel = self._create_tab_panel()
        splitter.addWidget(right_panel)

        # 設置分割比例
        splitter.setSizes([320, 1080])
        splitter.setStretchFactor(0, 0)  # 左側固定
        splitter.setStretchFactor(1, 1)  # 右側可伸縮
        splitter.setCollapsible(0, False)
        splitter.setCollapsible(1, False)

        main_layout.addWidget(splitter)

    def _create_control_panel(self) -> QWidget:
        """建立左側控制面板（含捲動支援）"""
        from PySide6.QtWidgets import (
            QGroupBox,
            QFormLayout,
            QPushButton,
            QSpinBox,
            QDoubleSpinBox,
            QListWidget,
            QScrollArea,
        )

        # 外層容器
        container = QFrame()
        container.setObjectName("controlPanel")
        container.setMinimumWidth(260)
        container.setMaximumWidth(380)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # 捲動區域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.NoFrame)

        # 內容面板
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # ===== 標題 =====
        title_label = QLabel("相機標定工具")
        title_label.setProperty("heading", True)
        layout.addWidget(title_label)

        subtitle_label = QLabel("使用棋盤格圖案進行相機校正")
        subtitle_label.setProperty("subheading", True)
        subtitle_label.setWordWrap(True)
        layout.addWidget(subtitle_label)

        layout.addSpacing(4)

        # ===== 棋盤格設定 =====
        cb_group = QGroupBox("棋盤格參數")
        cb_layout = QFormLayout(cb_group)
        cb_layout.setSpacing(8)
        cb_layout.setContentsMargins(12, 20, 12, 12)

        self.cols_spin = QSpinBox()
        self.cols_spin.setRange(2, 50)
        self.cols_spin.setValue(17)
        self.cols_spin.setToolTip("棋盤格內部角點的水平數量（與原工具「寬度」相同）")
        cb_layout.addRow("寬度(角點數)：", self.cols_spin)

        self.rows_spin = QSpinBox()
        self.rows_spin.setRange(2, 50)
        self.rows_spin.setValue(12)
        self.rows_spin.setToolTip("棋盤格內部角點的垂直數量（與原工具「高度」相同）")
        cb_layout.addRow("高度(角點數)：", self.rows_spin)

        self.square_size_spin = QDoubleSpinBox()
        self.square_size_spin.setRange(0.1, 100.0)
        self.square_size_spin.setValue(1.0)
        self.square_size_spin.setDecimals(2)
        self.square_size_spin.setSuffix(" cm")
        self.square_size_spin.setToolTip("棋盤格每個方格的實際邊長（公分）")
        cb_layout.addRow("方格邊長：", self.square_size_spin)

        layout.addWidget(cb_group)

        # ===== 圖像列表 =====
        img_group = QGroupBox("標定圖像")
        img_layout = QVBoxLayout(img_group)
        img_layout.setContentsMargins(12, 20, 12, 12)
        img_layout.setSpacing(8)

        self.image_list = QListWidget()
        self.image_list.setMinimumHeight(100)
        self.image_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.image_list.setToolTip("已載入的標定圖像列表，點擊可預覽")
        self.image_list.currentItemChanged.connect(self._on_image_selected)
        img_layout.addWidget(self.image_list)

        # 圖像操作按鈕
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self.add_images_btn = QPushButton("新增圖像")
        self.add_images_btn.clicked.connect(self._on_add_images)
        self.add_images_btn.setToolTip("選擇標定用的圖像檔案")
        btn_layout.addWidget(self.add_images_btn)

        self.clear_images_btn = QPushButton("清除全部")
        self.clear_images_btn.setProperty("secondary", True)
        self.clear_images_btn.clicked.connect(self._on_clear_images)
        self.clear_images_btn.setToolTip("移除所有已載入的圖像")
        btn_layout.addWidget(self.clear_images_btn)

        img_layout.addLayout(btn_layout)
        layout.addWidget(img_group, 1)  # stretch factor 1

        # ===== 操作按鈕 =====
        action_group = QGroupBox("執行操作")
        action_layout = QVBoxLayout(action_group)
        action_layout.setContentsMargins(12, 20, 12, 12)
        action_layout.setSpacing(8)

        self.detect_btn = QPushButton("偵測角點")
        self.detect_btn.clicked.connect(self._on_detect_corners)
        self.detect_btn.setToolTip("在所有圖像中偵測棋盤格角點")
        action_layout.addWidget(self.detect_btn)

        self.calibrate_btn = QPushButton("執行標定")
        self.calibrate_btn.clicked.connect(self._on_calibrate)
        self.calibrate_btn.setToolTip("計算相機內參矩陣與畸變係數")
        action_layout.addWidget(self.calibrate_btn)

        self.export_btn = QPushButton("匯出結果")
        self.export_btn.clicked.connect(self._on_export)
        self.export_btn.setEnabled(False)
        self.export_btn.setToolTip("將標定結果儲存為檔案")
        action_layout.addWidget(self.export_btn)

        layout.addWidget(action_group)

        # ===== 進度條 =====
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        layout.addWidget(self.progress_bar)

        # 設置捲動區域
        scroll.setWidget(panel)
        container_layout.addWidget(scroll)

        return container

    def _create_tab_panel(self) -> QWidget:
        """建立右側標籤頁面板"""
        panel = QFrame()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)

        self.tab_widget = QTabWidget()
        self.tab_widget.addTab(self._create_webcam_tab(), "WebCAM")
        self.tab_widget.addTab(self._create_intrinsic_tab(), "內參標定")
        self.tab_widget.addTab(self._create_points_tab(), "點位數據")
        self.tab_widget.addTab(self._create_extrinsic_tab(), "外參計算")
        self.tab_widget.addTab(self._create_transform_tab(), "座標轉換")

        layout.addWidget(self.tab_widget)
        return panel

    def _create_webcam_tab(self) -> QWidget:
        """建立 WebCAM 頁籤"""
        from PySide6.QtWidgets import (
            QGroupBox, QPushButton, QComboBox, QFormLayout, QTableWidget,
            QTableWidgetItem, QHeaderView,
        )
        from PySide6.QtCore import QTimer

        container = QWidget()
        main_layout = QVBoxLayout(container)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # ========== 上方：控制區 ==========
        ctrl_container = QWidget()
        ctrl_layout = QHBoxLayout(ctrl_container)
        ctrl_layout.setContentsMargins(0, 0, 0, 0)
        ctrl_layout.setSpacing(16)

        # --- 相機選擇 ---
        cam_group = QGroupBox("相機選擇")
        cam_layout = QFormLayout(cam_group)
        cam_layout.setContentsMargins(12, 16, 12, 12)
        cam_layout.setSpacing(8)

        cam_select_layout = QHBoxLayout()
        cam_select_layout.setSpacing(8)

        self.cam_combo = QComboBox()
        self.cam_combo.setMinimumWidth(200)
        self.cam_combo.setToolTip("選擇要使用的相機裝置")
        cam_select_layout.addWidget(self.cam_combo)

        self.refresh_cam_btn = QPushButton("掃描")
        self.refresh_cam_btn.setFixedWidth(80)
        self.refresh_cam_btn.clicked.connect(self._scan_cameras)
        self.refresh_cam_btn.setToolTip("重新掃描可用的相機裝置")
        cam_select_layout.addWidget(self.refresh_cam_btn)

        cam_layout.addRow("裝置：", cam_select_layout)
        ctrl_layout.addWidget(cam_group)

        # --- 串流控制 ---
        stream_group = QGroupBox("串流控制")
        stream_layout = QHBoxLayout(stream_group)
        stream_layout.setContentsMargins(12, 16, 12, 12)
        stream_layout.setSpacing(8)

        self.start_stream_btn = QPushButton("開始串流")
        self.start_stream_btn.clicked.connect(self._start_stream)
        self.start_stream_btn.setToolTip("開始接收相機影像")
        stream_layout.addWidget(self.start_stream_btn)

        self.stop_stream_btn = QPushButton("停止串流")
        self.stop_stream_btn.clicked.connect(self._stop_stream)
        self.stop_stream_btn.setEnabled(False)
        self.stop_stream_btn.setProperty("secondary", True)
        self.stop_stream_btn.setToolTip("停止接收相機影像")
        stream_layout.addWidget(self.stop_stream_btn)

        ctrl_layout.addWidget(stream_group)

        # --- 拍照控制 ---
        capture_group = QGroupBox("拍照")
        capture_layout = QHBoxLayout(capture_group)
        capture_layout.setContentsMargins(12, 16, 12, 12)
        capture_layout.setSpacing(8)

        self.capture_btn = QPushButton("拍照")
        self.capture_btn.clicked.connect(self._capture_photo)
        self.capture_btn.setEnabled(False)
        self.capture_btn.setToolTip("擷取當前畫面並暫停串流")
        capture_layout.addWidget(self.capture_btn)

        self.resume_btn = QPushButton("繼續串流")
        self.resume_btn.clicked.connect(self._resume_stream)
        self.resume_btn.setEnabled(False)
        self.resume_btn.setProperty("secondary", True)
        self.resume_btn.setToolTip("繼續接收相機影像")
        capture_layout.addWidget(self.resume_btn)

        self.save_photo_btn = QPushButton("儲存照片")
        self.save_photo_btn.clicked.connect(self._save_captured_photo)
        self.save_photo_btn.setEnabled(False)
        self.save_photo_btn.setProperty("secondary", True)
        self.save_photo_btn.setToolTip("將拍攝的照片儲存到檔案")
        capture_layout.addWidget(self.save_photo_btn)

        ctrl_layout.addWidget(capture_group)
        ctrl_layout.addStretch()

        main_layout.addWidget(ctrl_container)

        # ========== 中央：左右分割 - 影像 + 標記點列表 ==========
        center_splitter = QSplitter(Qt.Horizontal)

        # --- 左側：影像顯示區 ---
        image_container = QWidget()
        image_layout = QVBoxLayout(image_container)
        image_layout.setContentsMargins(0, 0, 0, 0)
        image_layout.setSpacing(4)

        # 使用自定義的 ImageViewer（支持縮放、拖曳）
        self.image_viewer = ImageViewer()
        self.image_viewer.point_added.connect(self._on_point_added)
        self.image_viewer.point_selected.connect(self._on_point_selected)
        self.image_viewer.mouse_moved.connect(self._on_viewer_mouse_moved)
        self.image_viewer.mouse_left.connect(self._on_viewer_mouse_left)
        image_layout.addWidget(self.image_viewer, 1)

        # 座標顯示列
        coord_layout = QHBoxLayout()
        self.webcam_status = QLabel("尚未連接相機")
        self.webcam_status.setProperty("subheading", True)
        coord_layout.addWidget(self.webcam_status)
        coord_layout.addStretch()

        self.webcam_coord_label = QLabel("")
        self.webcam_coord_label.setProperty("subheading", True)
        self.webcam_coord_label.setStyleSheet("color: #1a73e8; font-weight: bold;")
        coord_layout.addWidget(self.webcam_coord_label)

        self.webcam_resolution = QLabel("")
        self.webcam_resolution.setProperty("subheading", True)
        coord_layout.addWidget(self.webcam_resolution)

        image_layout.addLayout(coord_layout)

        center_splitter.addWidget(image_container)

        # --- 右側：標記點列表 ---
        points_container = QWidget()
        points_layout = QVBoxLayout(points_container)
        points_layout.setContentsMargins(0, 0, 0, 0)
        points_layout.setSpacing(8)

        points_header = QLabel("標記點 (點擊圖片新增)")
        points_header.setProperty("subheading", True)
        points_header.setStyleSheet("font-weight: bold;")
        points_layout.addWidget(points_header)

        self.marked_points_table = QTableWidget()
        self.marked_points_table.setColumnCount(3)
        self.marked_points_table.setHorizontalHeaderLabels(["名稱", "X (px)", "Y (px)"])
        self.marked_points_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.marked_points_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.marked_points_table.setSelectionMode(QTableWidget.SingleSelection)
        # 允許雙擊編輯（只有名稱欄位可編輯）
        self.marked_points_table.setEditTriggers(QTableWidget.DoubleClicked)
        self.marked_points_table.cellChanged.connect(self._on_point_name_changed)
        self.marked_points_table.setMinimumWidth(200)
        self.marked_points_table.setMaximumWidth(300)
        self.marked_points_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dadce0;
                border-radius: 4px;
                background-color: #ffffff;
                gridline-color: #e8eaed;
            }
            QTableWidget::item {
                padding: 4px 8px;
            }
            QTableWidget::item:selected {
                background-color: #1a73e8;
                color: #ffffff;
            }
            QTableWidget::item:hover:!selected {
                background-color: #e8f0fe;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                border: none;
                border-bottom: 1px solid #dadce0;
                padding: 6px 8px;
                font-weight: bold;
            }
        """)
        # 安裝事件過濾器處理鍵盤刪除
        self.marked_points_table.installEventFilter(self)
        points_layout.addWidget(self.marked_points_table, 1)

        # 標記點操作按鈕
        points_btn_layout = QHBoxLayout()
        points_btn_layout.setSpacing(4)

        self.delete_point_btn = QPushButton("刪除選中")
        self.delete_point_btn.clicked.connect(self._delete_selected_point)
        self.delete_point_btn.setProperty("secondary", True)
        points_btn_layout.addWidget(self.delete_point_btn)

        self.clear_points_btn = QPushButton("清除全部")
        self.clear_points_btn.clicked.connect(self._clear_all_points)
        self.clear_points_btn.setProperty("secondary", True)
        points_btn_layout.addWidget(self.clear_points_btn)

        points_layout.addLayout(points_btn_layout)

        # 提示
        points_hint = QLabel("拍照後點擊圖片標記點位\n點位將用於外參計算")
        points_hint.setProperty("subheading", True)
        points_hint.setStyleSheet("color: #5f6368; font-size: 11px;")
        points_hint.setWordWrap(True)
        points_layout.addWidget(points_hint)

        center_splitter.addWidget(points_container)
        center_splitter.setSizes([700, 250])
        center_splitter.setStretchFactor(0, 1)
        center_splitter.setStretchFactor(1, 0)

        main_layout.addWidget(center_splitter, 1)

        # ===== WebCAM 狀態變數 =====
        self._webcam_worker = None
        self._captured_frame = None
        self._is_paused = False
        self._current_frame = None

        # 影像尺寸
        self._image_width = 0
        self._image_height = 0

        # 初始掃描相機
        self._scan_cameras()

        return container

    # ===== ImageViewer 信號處理 =====

    def _on_point_added(self, img_x: float, img_y: float):
        """新增標記點"""
        # 生成點名稱 (A, B, C, ..., AA, AB, ...)
        point_index = len(self._marked_points)
        if point_index < 26:
            point_name = chr(ord('A') + point_index)
        else:
            point_name = chr(ord('A') + point_index // 26 - 1) + chr(ord('A') + point_index % 26)

        # 新增標記點
        self._marked_points.append({
            'name': point_name,
            'pixel_x': img_x,
            'pixel_y': img_y,
            'world_x': None,
            'world_y': None,
        })

        # 更新顯示
        self._refresh_points_display(select_row=len(self._marked_points) - 1)
        self.webcam_status.setText(f"已標記點 {point_name} ({img_x:.1f}, {img_y:.1f})")

    def _on_point_selected(self, index: int):
        """選中現有標記點"""
        if 0 <= index < len(self._marked_points):
            self._select_table_row(index)
            point = self._marked_points[index]
            self.webcam_status.setText(f"已選中點 {point['name']}")

    def _on_viewer_mouse_moved(self, img_x: float, img_y: float):
        """滑鼠在圖像上移動"""
        self.webcam_coord_label.setText(f"({img_x:.1f}, {img_y:.1f})")

    def _on_viewer_mouse_left(self):
        """滑鼠離開圖像"""
        self.webcam_coord_label.setText("")

    def eventFilter(self, obj, event):
        """事件過濾器 - 處理表格鍵盤事件"""
        from PySide6.QtCore import QEvent
        from PySide6.QtGui import QKeyEvent

        if obj == self.marked_points_table and event.type() == QEvent.KeyPress:
            key = event.key()
            # Delete 或 Backspace 刪除選中點
            if key in (Qt.Key_Delete, Qt.Key_Backspace):
                self._delete_selected_point()
                return True

        return super().eventFilter(obj, event)

    def _on_point_name_changed(self, row: int, column: int):
        """當標記點名稱被編輯時"""
        import re

        # 只處理名稱欄位（第 0 欄）
        if column != 0 or row >= len(self._marked_points):
            return

        item = self.marked_points_table.item(row, column)
        if item is None:
            return

        new_name = item.text().strip()
        old_name = self._marked_points[row]['name']

        # 驗證名稱
        # 1. 移除中文字符（只保留英文、數字、底線、連字號）
        new_name = re.sub(r'[^\x00-\x7F]', '', new_name)
        # 2. 限制 10 個字元
        new_name = new_name[:10]
        # 3. 如果為空，恢復舊名稱
        if not new_name:
            new_name = old_name

        # 更新數據和顯示
        self._marked_points[row]['name'] = new_name

        # 如果名稱被修改過，更新表格顯示
        if new_name != item.text():
            self.marked_points_table.blockSignals(True)
            item.setText(new_name)
            self.marked_points_table.blockSignals(False)

        # 更新 ImageViewer 中的標記點顯示
        self.image_viewer.set_points(self._marked_points)

        # 更新外參表格
        self._update_extrinsic_points_table()

        self.webcam_status.setText(f"點位名稱已更新為 {new_name}")

    def _refresh_points_display(self, select_row: int = -1):
        """刷新標記點顯示（表格 + 圖像）"""
        # 更新表格
        self._update_marked_points_table()

        # 選中指定行
        if select_row >= 0:
            self._select_table_row(select_row)

        # 更新 ImageViewer 中的標記點
        self.image_viewer.set_points(self._marked_points)

        # 更新外參頁籤
        self._update_extrinsic_points_table()

    def _update_marked_points_table(self, select_last: bool = False):
        """更新標記點表格"""
        from PySide6.QtWidgets import QTableWidgetItem
        from PySide6.QtCore import QTimer

        # 暫時阻止信號避免觸發不必要的事件
        self.marked_points_table.blockSignals(True)

        self.marked_points_table.setRowCount(len(self._marked_points))
        for i, point in enumerate(self._marked_points):
            # 名稱欄位（可編輯）
            name_item = QTableWidgetItem(point['name'])
            self.marked_points_table.setItem(i, 0, name_item)

            # X 座標（唯讀）
            x_item = QTableWidgetItem(f"{point['pixel_x']:.1f}")
            x_item.setFlags(x_item.flags() & ~Qt.ItemIsEditable)
            self.marked_points_table.setItem(i, 1, x_item)

            # Y 座標（唯讀）
            y_item = QTableWidgetItem(f"{point['pixel_y']:.1f}")
            y_item.setFlags(y_item.flags() & ~Qt.ItemIsEditable)
            self.marked_points_table.setItem(i, 2, y_item)

        self.marked_points_table.blockSignals(False)

        # 選中最後一個項目（高亮新加的點）- 延遲執行確保 UI 更新完成
        if select_last and len(self._marked_points) > 0:
            last_row = len(self._marked_points) - 1
            # 使用 QTimer.singleShot 確保在事件循環中執行
            QTimer.singleShot(0, lambda: self._select_table_row(last_row))

    def _select_table_row(self, row: int):
        """選中表格中指定的行"""
        if row < self.marked_points_table.rowCount():
            self.marked_points_table.clearSelection()
            self.marked_points_table.selectRow(row)
            item = self.marked_points_table.item(row, 0)
            if item:
                self.marked_points_table.scrollToItem(item)

    def _delete_selected_point(self):
        """刪除選中的標記點"""
        selected_rows = self.marked_points_table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # 從後往前刪除，避免索引問題
        indices = sorted([row.row() for row in selected_rows], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self._marked_points):
                del self._marked_points[idx]

        # 重新命名點位
        self._rename_points()

        # 更新顯示（使用統一的刷新方法）
        self._refresh_points_display()
        self.webcam_status.setText(f"已刪除 {len(indices)} 個標記點")

    def _clear_all_points(self):
        """清除所有標記點"""
        if not self._marked_points:
            return

        reply = QMessageBox.question(
            self, "確認清除",
            f"確定要清除所有 {len(self._marked_points)} 個標記點嗎？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self._marked_points.clear()
            # 更新顯示（使用統一的刷新方法）
            self._refresh_points_display()
            self.webcam_status.setText("已清除所有標記點")

    def _rename_points(self):
        """重新命名所有點位（A, B, C, ...）"""
        for i, point in enumerate(self._marked_points):
            if i < 26:
                point['name'] = chr(ord('A') + i)
            else:
                point['name'] = chr(ord('A') + i // 26 - 1) + chr(ord('A') + i % 26)

    def _scan_cameras(self):
        """掃描可用的相機裝置"""
        import cv2

        self.cam_combo.clear()
        self.webcam_status.setText("正在掃描相機...")
        QApplication.processEvents()

        available_cameras = []
        for i in range(5):  # 檢查前 5 個索引（減少掃描時間）
            cap = cv2.VideoCapture(i, cv2.CAP_DSHOW)
            if cap.isOpened():
                # 先設置高解析度再讀取
                cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
                cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)

                ret, _ = cap.read()
                if ret:
                    # 獲取實際解析度
                    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    available_cameras.append((i, f"Camera {i} ({width}x{height})"))
                cap.release()

        if available_cameras:
            for idx, name in available_cameras:
                self.cam_combo.addItem(name, idx)
            self.webcam_status.setText(f"找到 {len(available_cameras)} 個相機裝置")
        else:
            self.cam_combo.addItem("未找到相機", -1)
            self.webcam_status.setText("未找到可用的相機裝置")

    def _start_stream(self):
        """開始相機串流（使用背景執行緒）"""
        cam_idx = self.cam_combo.currentData()
        if cam_idx is None or cam_idx < 0:
            QMessageBox.warning(self, "警告", "請先選擇有效的相機裝置")
            return

        # 創建並啟動 Worker
        self._webcam_worker = WebCamWorker(cam_index=cam_idx, parent=self)
        self._webcam_worker.preview_ready.connect(self._on_preview_received)
        self._webcam_worker.resolution_ready.connect(self._on_resolution_ready)
        self._webcam_worker.error.connect(self._on_webcam_error)
        self._webcam_worker.start()

        # 更新按鈕狀態
        self.start_stream_btn.setEnabled(False)
        self.stop_stream_btn.setEnabled(True)
        self.capture_btn.setEnabled(True)
        self.resume_btn.setEnabled(False)
        self.cam_combo.setEnabled(False)
        self.refresh_cam_btn.setEnabled(False)

        self._is_paused = False
        self._captured_frame = None
        self.webcam_status.setText("啟動中...")
        self.webcam_resolution.setText("")

    def _on_resolution_ready(self, width: int, height: int):
        """相機啟動後接收實際解析度"""
        self.webcam_resolution.setText(f"{width} x {height}")
        self.webcam_status.setText("串流中...")
        self._image_width = width
        self._image_height = height

    def _on_webcam_error(self, error_msg: str):
        """處理 WebCAM 錯誤"""
        QMessageBox.critical(self, "相機錯誤", error_msg)
        self._stop_stream()

    def _on_preview_received(self, preview_qimage: QImage):
        """
        接收背景執行緒傳來的預覽影像（在主執行緒中執行）

        Args:
            preview_qimage: 已縮放的預覽用 QImage
        """
        from PySide6.QtGui import QPixmap

        if self._is_paused:
            return

        # 將預覽圖縮放至填滿顯示區域
        pixmap = QPixmap.fromImage(preview_qimage)
        label_w = self.image_viewer.width()
        label_h = self.image_viewer.height()

        if label_w > 0 and label_h > 0:
            scaled = pixmap.scaled(
                label_w, label_h,
                Qt.KeepAspectRatio,
                Qt.FastTransformation  # 快速縮放
            )
            self.image_viewer.setPixmap(scaled)
        else:
            self.image_viewer.setPixmap(pixmap)

    def _stop_stream(self):
        """停止相機串流"""
        if self._webcam_worker is not None:
            self._webcam_worker.stop()
            self._webcam_worker = None

        # 更新按鈕狀態
        self.start_stream_btn.setEnabled(True)
        self.stop_stream_btn.setEnabled(False)
        self.capture_btn.setEnabled(False)
        self.resume_btn.setEnabled(False)
        self.save_photo_btn.setEnabled(False)
        self.cam_combo.setEnabled(True)
        self.refresh_cam_btn.setEnabled(True)

        self._is_paused = False
        self._captured_frame = None
        self.webcam_status.setText("串流已停止")
        self.webcam_resolution.setText("")

        # 清除 ImageViewer
        self.image_viewer.set_image(None, interactive=False)
        self.image_viewer.setText("請選擇相機並點擊「開始串流」")

    def _capture_photo(self):
        """拍照並暫停串流"""
        from PySide6.QtGui import QPixmap
        import cv2

        if self._webcam_worker is None:
            return

        # 從背景執行緒獲取原始幀（線程安全）
        frame, width, height = self._webcam_worker.get_frame()
        if frame is None:
            self.webcam_status.setText("拍照失敗：無法獲取畫面")
            return

        # 保存拍攝的照片（BGR 格式）
        self._captured_frame = frame
        self._is_paused = True
        self._image_width = width
        self._image_height = height

        # 轉換為 QPixmap 並顯示
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_frame.shape
        q_img = QImage(rgb_frame.data, w, h, ch * w, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_img)

        # 使用 ImageViewer 顯示（啟用交互模式）
        self.image_viewer.set_image(pixmap, interactive=True)
        self.image_viewer.set_points(self._marked_points)
        self.image_viewer.reset_view()

        # 更新按鈕狀態
        self.capture_btn.setEnabled(False)
        self.resume_btn.setEnabled(True)
        self.save_photo_btn.setEnabled(True)

        if self._marked_points:
            self.webcam_status.setText(f"已拍攝照片 - 點擊圖片標記點位 ({len(self._marked_points)} 個點)")
        else:
            self.webcam_status.setText("已拍攝照片 - 點擊圖片標記點位")

    def _resume_stream(self):
        """繼續串流"""
        self._is_paused = False
        self._captured_frame = None

        # 重置 ImageViewer 狀態（關閉交互模式，串流時不需要）
        self.image_viewer._interactive = False
        self.image_viewer._pixmap = None

        # 更新按鈕狀態
        self.capture_btn.setEnabled(True)
        self.resume_btn.setEnabled(False)
        self.save_photo_btn.setEnabled(False)
        self.webcam_status.setText("串流中...")

    def _save_captured_photo(self):
        """儲存拍攝的照片"""
        import cv2
        from datetime import datetime

        if self._captured_frame is None:
            return

        # 預設檔名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"capture_{timestamp}.jpg"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "儲存照片",
            default_name,
            "JPEG 圖像 (*.jpg);;PNG 圖像 (*.png);;所有檔案 (*.*)"
        )

        if file_path:
            cv2.imwrite(file_path, self._captured_frame)
            self.webcam_status.setText(f"照片已儲存: {Path(file_path).name}")

            # 詢問是否加入標定圖像列表
            reply = QMessageBox.question(
                self,
                "加入標定圖像",
                "是否將此照片加入標定圖像列表？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply == QMessageBox.Yes:
                self._add_image_to_list(file_path)

    def _add_image_to_list(self, file_path: str):
        """將圖像加入標定圖像列表"""
        self.image_list.addItem(Path(file_path).name)
        item = self.image_list.item(self.image_list.count() - 1)
        item.setData(Qt.UserRole, file_path)
        self.statusbar.showMessage(f"已加入圖像: {Path(file_path).name}")
        logger.info(f"加入圖像: {file_path}")

        # 如果已有內參，更新外參圖像下拉選單
        if self._result is not None:
            self._update_extrinsic_image_combo()

    def _update_extrinsic_points_table(self):
        """將 WebCAM 標記點同步到外參計算表格"""
        from PySide6.QtWidgets import QTableWidgetItem

        # 如果表格尚未創建，跳過
        if not hasattr(self, 'ext_webcam_points_table'):
            return

        # 暫時斷開信號以避免遞迴
        self.ext_webcam_points_table.blockSignals(True)

        # 更新表格
        self.ext_webcam_points_table.setRowCount(len(self._marked_points))

        for i, point in enumerate(self._marked_points):
            # 名稱欄位（唯讀）
            name_item = QTableWidgetItem(point['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.ext_webcam_points_table.setItem(i, 0, name_item)

            # 像素 X（唯讀）
            px_item = QTableWidgetItem(f"{point['pixel_x']:.1f}")
            px_item.setFlags(px_item.flags() & ~Qt.ItemIsEditable)
            self.ext_webcam_points_table.setItem(i, 1, px_item)

            # 像素 Y（唯讀）
            py_item = QTableWidgetItem(f"{point['pixel_y']:.1f}")
            py_item.setFlags(py_item.flags() & ~Qt.ItemIsEditable)
            self.ext_webcam_points_table.setItem(i, 2, py_item)

            # 世界 X（可編輯）
            wx_val = "" if point['world_x'] is None else f"{point['world_x']:.2f}"
            wx_item = QTableWidgetItem(wx_val)
            self.ext_webcam_points_table.setItem(i, 3, wx_item)

            # 世界 Y（可編輯）
            wy_val = "" if point['world_y'] is None else f"{point['world_y']:.2f}"
            wy_item = QTableWidgetItem(wy_val)
            self.ext_webcam_points_table.setItem(i, 4, wy_item)

        # 恢復信號
        self.ext_webcam_points_table.blockSignals(False)

        # 更新狀態和計算按鈕
        self._check_extrinsic_ready()

    def _on_ext_world_coord_changed(self, row: int, column: int):
        """當外參表格中的世界座標被修改時"""
        # 只處理世界座標欄位（第 3、4 欄）
        if column < 3 or row >= len(self._marked_points):
            return

        item = self.ext_webcam_points_table.item(row, column)
        if item is None:
            return

        text = item.text().strip()

        try:
            if text == "":
                value = None
            else:
                value = float(text)

            if column == 3:
                self._marked_points[row]['world_x'] = value
            elif column == 4:
                self._marked_points[row]['world_y'] = value

        except ValueError:
            # 無效輸入，恢復原值
            if column == 3:
                old_val = self._marked_points[row]['world_x']
            else:
                old_val = self._marked_points[row]['world_y']

            self.ext_webcam_points_table.blockSignals(True)
            item.setText("" if old_val is None else f"{old_val:.2f}")
            self.ext_webcam_points_table.blockSignals(False)
            return

        # 更新狀態
        self._check_extrinsic_ready()

    def _check_extrinsic_ready(self):
        """檢查外參計算是否滿足條件，更新按鈕和狀態"""
        if not hasattr(self, 'ext_webcam_points_table'):
            return

        # 計算有完整世界座標的點數
        valid_points = sum(
            1 for p in self._marked_points
            if p['world_x'] is not None and p['world_y'] is not None
        )

        # 獲取當前算法的最低點數要求
        algo = self.ext_algo_combo.currentData()
        if algo == "SOLVEPNP_P3P":
            min_points = 3
            algo_name = "P3P"
        elif algo == "SOLVEPNP_AP3P":
            min_points = 3
            algo_name = "AP3P"
        else:
            min_points = 4
            algo_name = "PnP"

        # 更新狀態標籤
        total_points = len(self._marked_points)
        if total_points == 0:
            self.ext_webcam_points_status.setText("尚無標記點（請在 WebCAM 頁籤拍照並標記）")
            self.ext_webcam_points_status.setStyleSheet("color: #ea4335;")
        elif valid_points < min_points:
            self.ext_webcam_points_status.setText(
                f"已標記 {total_points} 點，有效 {valid_points} 點（{algo_name} 需要 ≥{min_points} 點）"
            )
            self.ext_webcam_points_status.setStyleSheet("color: #ea4335;")
        else:
            self.ext_webcam_points_status.setText(
                f"已標記 {total_points} 點，有效 {valid_points} 點 ✓"
            )
            self.ext_webcam_points_status.setStyleSheet("color: #34a853;")

        # 更新計算按鈕狀態（需要內參和足夠的點）
        has_intrinsic = self._result is not None and self._result.intrinsic is not None
        can_calculate = has_intrinsic and valid_points >= min_points

        # 如果選擇使用 WebCAM 點位，則用這個條件
        if hasattr(self, 'ext_use_points_radio') and self.ext_use_points_radio.isChecked():
            self.ext_calibrate_btn.setEnabled(can_calculate)

    def _create_intrinsic_tab(self) -> QWidget:
        """建立內參標定頁籤"""
        from PySide6.QtWidgets import QTextEdit, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem
        from PySide6.QtGui import QPixmap

        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # 上下分割
        splitter = QSplitter(Qt.Vertical)

        # ===== 上方：圖像預覽 =====
        preview_container = QFrame()
        preview_layout = QVBoxLayout(preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.setSpacing(4)

        preview_header = QLabel("圖像預覽")
        preview_header.setProperty("subheading", True)
        preview_layout.addWidget(preview_header)

        # 圖像檢視器
        self.image_scene = QGraphicsScene()
        self.image_view = QGraphicsView(self.image_scene)
        self.image_view.setMinimumHeight(200)
        self.image_view.setStyleSheet("""
            QGraphicsView {
                border: 1px solid #dadce0;
                border-radius: 8px;
                background-color: #f8f9fa;
            }
        """)
        from PySide6.QtGui import QPainter
        self.image_view.setRenderHints(
            QPainter.RenderHint.Antialiasing |
            QPainter.RenderHint.SmoothPixmapTransform
        )
        self.image_pixmap_item = None
        preview_layout.addWidget(self.image_view)

        self.image_info_label = QLabel("點擊左側列表中的圖像進行預覽")
        self.image_info_label.setProperty("subheading", True)
        self.image_info_label.setAlignment(Qt.AlignCenter)
        preview_layout.addWidget(self.image_info_label)

        splitter.addWidget(preview_container)

        # ===== 下方：標定結果 =====
        result_container = QFrame()
        result_layout = QVBoxLayout(result_container)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(4)

        result_header = QLabel("標定結果")
        result_header.setProperty("subheading", True)
        result_layout.addWidget(result_header)

        self.intrinsic_view = QTextEdit()
        self.intrinsic_view.setReadOnly(True)
        self.intrinsic_view.setPlaceholderText(
            "請依照以下步驟進行相機標定：\n\n"
            "① 設定棋盤格參數（行數、列數、方格邊長）\n"
            "② 點擊「新增圖像」載入標定照片\n"
            "③ 點擊「偵測角點」自動識別\n"
            "④ 點擊「執行標定」計算參數\n"
            "⑤ 點擊「匯出結果」儲存"
        )
        result_layout.addWidget(self.intrinsic_view)

        splitter.addWidget(result_container)

        # 設置分割比例 (60% 圖像, 40% 結果)
        splitter.setSizes([400, 250])

        layout.addWidget(splitter)
        return widget

    def _create_points_tab(self) -> QWidget:
        """建立點位數據管理頁籤"""
        from PySide6.QtWidgets import (
            QTextEdit, QGroupBox, QPushButton,
            QFormLayout, QTableWidget, QTableWidgetItem, QHeaderView,
            QDoubleSpinBox, QSpinBox, QScrollArea,
        )

        # 外層容器
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # 捲動區域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # 說明
        header = QLabel("點位數據管理")
        header.setProperty("heading", True)
        layout.addWidget(header)

        desc = QLabel("管理像素座標與世界座標的對應關係，用於外參計算")
        desc.setProperty("subheading", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # ===== 手動添加點位 =====
        add_group = QGroupBox("添加點位")
        add_layout = QFormLayout(add_group)
        add_layout.setContentsMargins(12, 20, 12, 12)
        add_layout.setSpacing(8)
        add_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        # 點位 ID
        self.point_id_spin = QSpinBox()
        self.point_id_spin.setRange(1, 9999)
        self.point_id_spin.setValue(1)
        self.point_id_spin.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        add_layout.addRow("點位 ID：", self.point_id_spin)

        # 像素座標
        pixel_layout = QHBoxLayout()
        pixel_layout.setSpacing(8)
        self.point_img_x_spin = QDoubleSpinBox()
        self.point_img_x_spin.setRange(0, 10000)
        self.point_img_x_spin.setDecimals(2)
        self.point_img_x_spin.setSuffix(" px")
        self.point_img_x_spin.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        pixel_layout.addWidget(QLabel("X:"))
        pixel_layout.addWidget(self.point_img_x_spin, 1)
        self.point_img_y_spin = QDoubleSpinBox()
        self.point_img_y_spin.setRange(0, 10000)
        self.point_img_y_spin.setDecimals(2)
        self.point_img_y_spin.setSuffix(" px")
        self.point_img_y_spin.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        pixel_layout.addWidget(QLabel("Y:"))
        pixel_layout.addWidget(self.point_img_y_spin, 1)
        add_layout.addRow("像素座標：", pixel_layout)

        # 世界座標
        world_layout = QHBoxLayout()
        world_layout.setSpacing(8)
        self.point_world_x_spin = QDoubleSpinBox()
        self.point_world_x_spin.setRange(-10000, 10000)
        self.point_world_x_spin.setDecimals(2)
        self.point_world_x_spin.setSuffix(" mm")
        self.point_world_x_spin.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        world_layout.addWidget(QLabel("X:"))
        world_layout.addWidget(self.point_world_x_spin, 1)
        self.point_world_y_spin = QDoubleSpinBox()
        self.point_world_y_spin.setRange(-10000, 10000)
        self.point_world_y_spin.setDecimals(2)
        self.point_world_y_spin.setSuffix(" mm")
        self.point_world_y_spin.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        world_layout.addWidget(QLabel("Y:"))
        world_layout.addWidget(self.point_world_y_spin, 1)
        add_layout.addRow("世界座標：", world_layout)

        # 添加按鈕
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)
        add_point_btn = QPushButton("添加點位")
        add_point_btn.clicked.connect(self._on_add_point)
        add_point_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_layout.addWidget(add_point_btn)

        clear_points_btn = QPushButton("清除全部")
        clear_points_btn.setProperty("secondary", True)
        clear_points_btn.clicked.connect(self._on_clear_points)
        clear_points_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_layout.addWidget(clear_points_btn)
        add_layout.addRow("", btn_layout)

        layout.addWidget(add_group)

        # ===== 點位列表 =====
        list_group = QGroupBox(f"當前點位 (共 0 個)")
        self.points_list_group = list_group
        list_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        list_layout = QVBoxLayout(list_group)
        list_layout.setContentsMargins(12, 20, 12, 12)

        # 表格
        self.points_table = QTableWidget()
        self.points_table.setColumnCount(5)
        self.points_table.setHorizontalHeaderLabels(["ID", "像素X", "像素Y", "世界X", "世界Y"])
        self.points_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.points_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.points_table.setMinimumHeight(150)
        list_layout.addWidget(self.points_table)

        # 匯入/匯出
        io_layout = QHBoxLayout()
        io_layout.setSpacing(8)
        import_csv_btn = QPushButton("匯入CSV")
        import_csv_btn.clicked.connect(self._on_import_points_csv)
        import_csv_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        io_layout.addWidget(import_csv_btn)

        export_csv_btn = QPushButton("匯出CSV")
        export_csv_btn.clicked.connect(self._on_export_points_csv)
        export_csv_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        io_layout.addWidget(export_csv_btn)

        load_from_corners_btn = QPushButton("從角點載入")
        load_from_corners_btn.clicked.connect(self._on_load_corners_to_points)
        load_from_corners_btn.setToolTip("將偵測到的角點載入為像素座標")
        load_from_corners_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        io_layout.addWidget(load_from_corners_btn)

        list_layout.addLayout(io_layout)
        layout.addWidget(list_group, 1)

        scroll.setWidget(widget)
        container_layout.addWidget(scroll)

        return container


    def _create_extrinsic_tab(self) -> QWidget:
        """建立外參計算頁籤"""
        from PySide6.QtWidgets import (
            QTextEdit, QGroupBox, QComboBox, QPushButton,
            QFormLayout, QRadioButton, QButtonGroup, QScrollArea,
        )

        # 外層容器
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # 捲動區域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # 說明
        header = QLabel("外參計算")
        header.setProperty("heading", True)
        layout.addWidget(header)

        desc = QLabel("使用 solvePnP 計算相機相對於世界座標系的位姿")
        desc.setProperty("subheading", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # ===== 計算方式選擇 =====
        method_group = QGroupBox("計算方式")
        method_layout = QVBoxLayout(method_group)
        method_layout.setContentsMargins(12, 20, 12, 12)

        self.ext_method_group = QButtonGroup(self)

        # 方式一：使用點位數據
        self.ext_use_points_radio = QRadioButton("使用點位數據（推薦）")
        self.ext_use_points_radio.setChecked(True)
        self.ext_use_points_radio.setToolTip("使用「點位數據」頁籤中的像素-世界座標對應")
        self.ext_method_group.addButton(self.ext_use_points_radio, 0)
        method_layout.addWidget(self.ext_use_points_radio)

        self.ext_points_status = QLabel("點位數據：0 個點")
        self.ext_points_status.setProperty("subheading", True)
        self.ext_points_status.setStyleSheet("margin-left: 24px;")
        method_layout.addWidget(self.ext_points_status)

        # 方式二：使用圖像角點
        self.ext_use_corners_radio = QRadioButton("使用圖像角點")
        self.ext_use_corners_radio.setToolTip("選擇一張圖像，使用棋盤格角點計算")
        self.ext_method_group.addButton(self.ext_use_corners_radio, 1)
        method_layout.addWidget(self.ext_use_corners_radio)

        # 圖像選擇下拉選單
        corners_layout = QHBoxLayout()
        corners_layout.setContentsMargins(24, 0, 0, 0)
        self.ext_image_combo = QComboBox()
        self.ext_image_combo.setToolTip("選擇要用於外參計算的圖像")
        self.ext_image_combo.addItem("-- 請先載入圖像 --")
        corners_layout.addWidget(QLabel("定位圖像："))
        corners_layout.addWidget(self.ext_image_combo, 1)
        method_layout.addLayout(corners_layout)

        layout.addWidget(method_group)

        # ===== WebCAM 標記點 =====
        webcam_points_group = QGroupBox("WebCAM 標記點")
        webcam_points_layout = QVBoxLayout(webcam_points_group)
        webcam_points_layout.setContentsMargins(12, 20, 12, 12)
        webcam_points_layout.setSpacing(8)

        webcam_points_desc = QLabel("從 WebCAM 頁籤標記的點位，請輸入對應的世界座標")
        webcam_points_desc.setProperty("subheading", True)
        webcam_points_desc.setWordWrap(True)
        webcam_points_layout.addWidget(webcam_points_desc)

        # 標記點表格
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
        self.ext_webcam_points_table = QTableWidget()
        self.ext_webcam_points_table.setColumnCount(5)
        self.ext_webcam_points_table.setHorizontalHeaderLabels([
            "名稱", "像素 X", "像素 Y", "世界 X", "世界 Y"
        ])
        self.ext_webcam_points_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ext_webcam_points_table.setMinimumHeight(120)
        self.ext_webcam_points_table.setMaximumHeight(200)
        self.ext_webcam_points_table.cellChanged.connect(self._on_ext_world_coord_changed)
        webcam_points_layout.addWidget(self.ext_webcam_points_table)

        # 點位狀態
        self.ext_webcam_points_status = QLabel("尚無標記點（請在 WebCAM 頁籤拍照並標記）")
        self.ext_webcam_points_status.setProperty("subheading", True)
        self.ext_webcam_points_status.setStyleSheet("color: #ea4335;")
        webcam_points_layout.addWidget(self.ext_webcam_points_status)

        layout.addWidget(webcam_points_group)

        # ===== Excel 數據管理 =====
        excel_group = QGroupBox("Excel 數據管理")
        excel_layout = QVBoxLayout(excel_group)
        excel_layout.setContentsMargins(12, 20, 12, 12)
        excel_layout.setSpacing(8)

        excel_desc = QLabel("匯出範本填入視覺座標與機械臂座標，再匯入計算外參")
        excel_desc.setProperty("subheading", True)
        excel_desc.setWordWrap(True)
        excel_layout.addWidget(excel_desc)

        excel_btn_layout = QHBoxLayout()
        excel_btn_layout.setSpacing(8)

        self.ext_export_template_btn = QPushButton("匯出Excel範本")
        self.ext_export_template_btn.setToolTip("匯出空白Excel範本，包含像素座標和機械臂座標欄位")
        self.ext_export_template_btn.clicked.connect(self._on_export_excel_template)
        excel_btn_layout.addWidget(self.ext_export_template_btn)

        self.ext_import_excel_btn = QPushButton("匯入Excel數據")
        self.ext_import_excel_btn.setToolTip("匯入填寫好的Excel數據")
        self.ext_import_excel_btn.clicked.connect(self._on_import_excel_data)
        excel_btn_layout.addWidget(self.ext_import_excel_btn)

        excel_layout.addLayout(excel_btn_layout)

        self.excel_data_status = QLabel("尚未匯入Excel數據")
        self.excel_data_status.setProperty("subheading", True)
        self.excel_data_status.setStyleSheet("margin-top: 4px;")
        excel_layout.addWidget(self.excel_data_status)

        layout.addWidget(excel_group)

        # ===== 算法選擇 =====
        algo_group = QGroupBox("算法選擇")
        algo_layout = QVBoxLayout(algo_group)
        algo_layout.setContentsMargins(12, 20, 12, 12)
        algo_layout.setSpacing(8)

        algo_select_layout = QFormLayout()
        self.ext_algo_combo = QComboBox()
        self.ext_algo_combo.addItem("PnP 迭代法 (≥4點, 推薦)", "SOLVEPNP_ITERATIVE")
        self.ext_algo_combo.addItem("EPnP 算法 (≥4點, 高效)", "SOLVEPNP_EPNP")
        self.ext_algo_combo.addItem("P3P 算法 (=3點)", "SOLVEPNP_P3P")
        self.ext_algo_combo.addItem("AP3P 算法 (≥3點)", "SOLVEPNP_AP3P")
        self.ext_algo_combo.addItem("IPPE 算法 (≥4點, 平面)", "SOLVEPNP_IPPE")
        self.ext_algo_combo.addItem("IPPE_SQUARE (≥4點, 正方形)", "SOLVEPNP_IPPE_SQUARE")
        self.ext_algo_combo.currentIndexChanged.connect(self._on_algo_changed)
        algo_select_layout.addRow("PnP 算法：", self.ext_algo_combo)
        algo_layout.addLayout(algo_select_layout)

        # 算法說明
        self.algo_desc_label = QLabel(
            "迭代法：最通用，適合大多數情況，需≥4個點"
        )
        self.algo_desc_label.setProperty("subheading", True)
        self.algo_desc_label.setWordWrap(True)
        self.algo_desc_label.setStyleSheet("color: #5f6368; font-size: 12px;")
        algo_layout.addWidget(self.algo_desc_label)

        layout.addWidget(algo_group)

        # ===== 執行按鈕 =====
        btn_layout = QHBoxLayout()

        self.ext_calibrate_btn = QPushButton("計算外參")
        self.ext_calibrate_btn.clicked.connect(self._on_calibrate_extrinsic)
        btn_layout.addWidget(self.ext_calibrate_btn)

        self.ext_export_btn = QPushButton("匯出外參")
        self.ext_export_btn.clicked.connect(self._on_export_extrinsic)
        self.ext_export_btn.setEnabled(False)
        btn_layout.addWidget(self.ext_export_btn)

        layout.addLayout(btn_layout)

        # ===== 結果顯示區 =====
        result_group = QGroupBox("外參結果")
        result_layout = QVBoxLayout(result_group)
        result_layout.setContentsMargins(12, 20, 12, 12)

        self.extrinsic_view = QTextEdit()
        self.extrinsic_view.setReadOnly(True)
        self.extrinsic_view.setPlaceholderText(
            "外參標定流程：\n\n"
            "方式一（推薦）：使用點位數據\n"
            "① 在「內參標定」完成內參計算或載入\n"
            "② 在「點位數據」設定像素-世界座標對應\n"
            "③ 在此頁選擇「使用點位數據」\n"
            "④ 點擊「計算外參」\n\n"
            "方式二：使用圖像角點\n"
            "① 在「內參標定」完成內參計算\n"
            "② 選擇一張圖像定義世界座標系\n"
            "③ 在此頁選擇「使用圖像角點」\n"
            "④ 點擊「計算外參」"
        )
        result_layout.addWidget(self.extrinsic_view)

        layout.addWidget(result_group, 1)

        scroll.setWidget(widget)
        container_layout.addWidget(scroll)

        return container

    def _create_transform_tab(self) -> QWidget:
        """建立座標轉換頁籤"""
        from PySide6.QtWidgets import (
            QTextEdit, QGroupBox, QDoubleSpinBox, QPushButton,
            QFormLayout, QGridLayout, QScrollArea,
        )

        # 外層容器
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # 捲動區域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        header = QLabel("座標轉換")
        header.setProperty("heading", True)
        layout.addWidget(header)

        # ===== 像素 → 世界 =====
        p2w_group = QGroupBox("像素 → 世界座標")
        p2w_layout = QFormLayout(p2w_group)
        p2w_layout.setContentsMargins(12, 20, 12, 12)
        p2w_layout.setSpacing(10)

        # 輸入像素座標
        pixel_input_layout = QHBoxLayout()
        self.pixel_u_spin = QDoubleSpinBox()
        self.pixel_u_spin.setRange(0, 10000)
        self.pixel_u_spin.setDecimals(1)
        self.pixel_u_spin.setSuffix(" px")
        pixel_input_layout.addWidget(QLabel("U:"))
        pixel_input_layout.addWidget(self.pixel_u_spin)

        self.pixel_v_spin = QDoubleSpinBox()
        self.pixel_v_spin.setRange(0, 10000)
        self.pixel_v_spin.setDecimals(1)
        self.pixel_v_spin.setSuffix(" px")
        pixel_input_layout.addWidget(QLabel("V:"))
        pixel_input_layout.addWidget(self.pixel_v_spin)
        p2w_layout.addRow("像素座標：", pixel_input_layout)

        # Z 平面
        self.world_z_spin = QDoubleSpinBox()
        self.world_z_spin.setRange(-10000, 10000)
        self.world_z_spin.setDecimals(2)
        self.world_z_spin.setValue(0)
        self.world_z_spin.setSuffix(" mm")
        self.world_z_spin.setToolTip("目標平面的 Z 座標（0 = 棋盤格平面）")
        p2w_layout.addRow("世界 Z 平面：", self.world_z_spin)

        # 轉換按鈕
        self.p2w_btn = QPushButton("轉換為世界座標")
        self.p2w_btn.clicked.connect(self._on_pixel_to_world)
        self.p2w_btn.setEnabled(False)
        p2w_layout.addRow("", self.p2w_btn)

        # 結果
        self.p2w_result = QLabel("X: -- mm, Y: -- mm")
        self.p2w_result.setStyleSheet("font-weight: bold; font-size: 16px; padding: 8px;")
        p2w_layout.addRow("結果：", self.p2w_result)

        layout.addWidget(p2w_group)

        # ===== 世界 → 像素 =====
        w2p_group = QGroupBox("世界 → 像素座標")
        w2p_layout = QFormLayout(w2p_group)
        w2p_layout.setContentsMargins(12, 20, 12, 12)
        w2p_layout.setSpacing(10)

        # 輸入世界座標
        world_input_layout = QHBoxLayout()
        self.world_x_spin = QDoubleSpinBox()
        self.world_x_spin.setRange(-10000, 10000)
        self.world_x_spin.setDecimals(2)
        self.world_x_spin.setSuffix(" mm")
        world_input_layout.addWidget(QLabel("X:"))
        world_input_layout.addWidget(self.world_x_spin)

        self.world_y_spin = QDoubleSpinBox()
        self.world_y_spin.setRange(-10000, 10000)
        self.world_y_spin.setDecimals(2)
        self.world_y_spin.setSuffix(" mm")
        world_input_layout.addWidget(QLabel("Y:"))
        world_input_layout.addWidget(self.world_y_spin)

        self.world_z_input_spin = QDoubleSpinBox()
        self.world_z_input_spin.setRange(-10000, 10000)
        self.world_z_input_spin.setDecimals(2)
        self.world_z_input_spin.setValue(0)
        self.world_z_input_spin.setSuffix(" mm")
        world_input_layout.addWidget(QLabel("Z:"))
        world_input_layout.addWidget(self.world_z_input_spin)
        w2p_layout.addRow("世界座標：", world_input_layout)

        # 轉換按鈕
        self.w2p_btn = QPushButton("轉換為像素座標")
        self.w2p_btn.clicked.connect(self._on_world_to_pixel)
        self.w2p_btn.setEnabled(False)
        w2p_layout.addRow("", self.w2p_btn)

        # 結果
        self.w2p_result = QLabel("U: -- px, V: -- px")
        self.w2p_result.setStyleSheet("font-weight: bold; font-size: 16px; padding: 8px;")
        w2p_layout.addRow("結果：", self.w2p_result)

        layout.addWidget(w2p_group)

        # 狀態提示
        self.transform_status = QLabel("請先完成內參和外參標定")
        self.transform_status.setProperty("subheading", True)
        self.transform_status.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.transform_status)

        layout.addStretch()

        scroll.setWidget(widget)
        container_layout.addWidget(scroll)

        return container

    def _setup_menu(self):
        """設置選單列"""
        menubar = self.menuBar()

        # ===== 檔案選單 =====
        file_menu = menubar.addMenu("檔案(&F)")

        open_action = QAction("開啟標定檔(&O)...", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._on_open)
        file_menu.addAction(open_action)

        save_action = QAction("儲存標定檔(&S)...", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self._on_export)
        file_menu.addAction(save_action)

        file_menu.addSeparator()

        exit_action = QAction("結束(&X)", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # ===== 檢視選單 =====
        view_menu = menubar.addMenu("檢視(&V)")

        self.theme_action = QAction("切換深色模式", self)
        self.theme_action.setShortcut("Ctrl+T")
        self.theme_action.triggered.connect(self._on_toggle_theme)
        view_menu.addAction(self.theme_action)

        # ===== 說明選單 =====
        help_menu = menubar.addMenu("說明(&H)")

        about_action = QAction("關於(&A)", self)
        about_action.triggered.connect(self._on_about)
        help_menu.addAction(about_action)

    def _setup_toolbar(self):
        """設置工具列"""
        toolbar = QToolBar("主工具列")
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(20, 20))
        self.addToolBar(toolbar)

        # 工具列按鈕
        add_action = toolbar.addAction("新增圖像")
        add_action.setToolTip("載入標定用圖像 (Ctrl+I)")
        add_action.triggered.connect(self._on_add_images)

        toolbar.addSeparator()

        detect_action = toolbar.addAction("偵測角點")
        detect_action.setToolTip("偵測棋盤格角點")
        detect_action.triggered.connect(self._on_detect_corners)

        calibrate_action = toolbar.addAction("執行標定")
        calibrate_action.setToolTip("計算相機參數")
        calibrate_action.triggered.connect(self._on_calibrate)

        toolbar.addSeparator()

        export_action = toolbar.addAction("匯出")
        export_action.setToolTip("匯出標定結果")
        export_action.triggered.connect(self._on_export)

        # 彈性空間 - 將主題指示器推到右側
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        toolbar.addWidget(spacer)

        # 主題指示器（僅顯示，從選單切換）
        self.theme_indicator = QLabel("🌙")
        self.theme_indicator.setToolTip("目前為明亮模式（從「檢視」選單切換）")
        self.theme_indicator.setStyleSheet("font-size: 18px; padding: 4px 12px;")
        toolbar.addWidget(self.theme_indicator)

    def _setup_statusbar(self):
        """設置狀態列"""
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        self.statusbar.showMessage("就緒")

    # ===== 事件處理 =====

    @Slot()
    def _on_toggle_theme(self):
        """切換主題"""
        self.theme_manager.toggle_theme()

        if self.theme_manager.is_dark:
            self.theme_indicator.setText("☀️")
            self.theme_indicator.setToolTip("目前為深色模式（從「檢視」選單切換）")
            self.theme_action.setText("切換明亮模式")
        else:
            self.theme_indicator.setText("🌙")
            self.theme_indicator.setToolTip("目前為明亮模式（從「檢視」選單切換）")
            self.theme_action.setText("切換深色模式")

    @Slot()
    def _on_add_images(self):
        """處理新增圖像"""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "選擇標定圖像",
            "",
            "圖像檔案 (*.jpg *.jpeg *.png *.bmp *.tiff);;所有檔案 (*)",
        )

        if files:
            for f in files:
                self.image_list.addItem(Path(f).name)
                item = self.image_list.item(self.image_list.count() - 1)
                item.setData(Qt.UserRole, f)

            self.statusbar.showMessage(f"已載入 {len(files)} 張圖像")
            logger.info(f"載入 {len(files)} 張圖像")

            # 如果已有內參，更新外參圖像下拉選單
            if self._result is not None:
                self._update_extrinsic_image_combo()

    @Slot()
    def _on_clear_images(self):
        """清除所有圖像"""
        self.image_list.clear()
        self._corner_cache.clear()
        self._clear_image_preview()
        self.statusbar.showMessage("已清除所有圖像")

    @Slot()
    def _on_image_selected(self, current, previous):
        """處理圖像選擇變更"""
        if current is None:
            self._clear_image_preview()
            return

        image_path = current.data(Qt.UserRole)
        if image_path:
            self._display_image(image_path)

    def _display_image(self, image_path: str):
        """顯示圖像（含角點標記）"""
        from PySide6.QtGui import QPixmap, QImage
        import cv2
        import numpy as np

        try:
            # 讀取圖像（支援中文路徑）
            with open(image_path, 'rb') as f:
                data = np.frombuffer(f.read(), dtype=np.uint8)
            img = cv2.imdecode(data, cv2.IMREAD_COLOR)

            if img is None:
                self.image_info_label.setText("無法載入圖像")
                return

            # 如果有角點資料，繪製角點
            if image_path in self._corner_cache:
                corners = self._corner_cache[image_path]
                if corners is not None:
                    cv2.drawChessboardCorners(
                        img,
                        (self.cols_spin.value(), self.rows_spin.value()),
                        corners,
                        True
                    )

            # 轉換為 QPixmap
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w, ch = img_rgb.shape
            bytes_per_line = ch * w
            q_img = QImage(img_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(q_img)

            # 顯示在場景中
            self.image_scene.clear()
            self.image_pixmap_item = self.image_scene.addPixmap(pixmap)
            self.image_scene.setSceneRect(pixmap.rect().toRectF())

            # 自適應縮放
            self.image_view.fitInView(
                self.image_scene.sceneRect(),
                Qt.KeepAspectRatio
            )

            # 更新資訊標籤
            filename = Path(image_path).name
            corner_status = "（已偵測角點）" if image_path in self._corner_cache else ""
            self.image_info_label.setText(f"{filename} - {w}×{h} {corner_status}")

        except Exception as e:
            self.image_info_label.setText(f"載入失敗：{e}")
            logger.error(f"載入圖像失敗：{e}")

    def _clear_image_preview(self):
        """清除圖像預覽"""
        self.image_scene.clear()
        self.image_pixmap_item = None
        self.image_info_label.setText("點擊左側列表中的圖像進行預覽")

    @Slot()
    def _on_detect_corners(self):
        """偵測角點（背景執行緒）"""
        if self.image_list.count() == 0:
            QMessageBox.warning(self, "提示", "請先載入圖像")
            return

        from vision_calib.core.types import CheckerboardConfig
        from vision_calib.utils.worker import CornerDetectionWorker

        config = CheckerboardConfig(
            rows=self.rows_spin.value(),
            cols=self.cols_spin.value(),
            square_size_mm=self.square_size_spin.value() * 10,  # cm → mm
        )

        # 取得圖像路徑
        paths = []
        for i in range(self.image_list.count()):
            item = self.image_list.item(i)
            paths.append(item.data(Qt.UserRole))

        # 禁用按鈕
        self._set_buttons_enabled(False)

        # 顯示進度條
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(paths))
        self.progress_bar.setValue(0)

        # 建立並啟動工作執行緒
        self._corner_worker = CornerDetectionWorker(paths, config, self)
        self._corner_worker.progress.connect(self._on_corner_progress)
        self._corner_worker.single_result.connect(self._on_corner_single_result)
        self._corner_worker.finished.connect(self._on_corner_finished)
        self._corner_worker.error.connect(self._on_corner_error)
        self._corner_worker.start()

    @Slot(int, int, str)
    def _on_corner_progress(self, current: int, total: int, message: str):
        """角點偵測進度更新"""
        self.progress_bar.setValue(current)
        self.statusbar.showMessage(message)

    @Slot(object)
    def _on_corner_single_result(self, result):
        """單張圖像角點偵測結果"""
        item = self.image_list.item(result.index)
        if item:
            filename = Path(result.image_path).name
            if result.success:
                item.setText(f"✓ {filename}")
                # 儲存角點到快取
                self._corner_cache[result.image_path] = result.corners
            else:
                item.setText(f"✗ {filename}")
                self._corner_cache[result.image_path] = None

            # 如果當前選中的是這張圖，更新預覽
            current_item = self.image_list.currentItem()
            if current_item and current_item.data(Qt.UserRole) == result.image_path:
                self._display_image(result.image_path)

    @Slot(int, int)
    def _on_corner_finished(self, success_count: int, total_count: int):
        """角點偵測完成"""
        self.progress_bar.setVisible(False)
        self._set_buttons_enabled(True)
        self.statusbar.showMessage(f"角點偵測完成：{success_count}/{total_count} 張成功")
        self._corner_worker = None

        # 如果已有內參數據，更新外參圖像下拉選單
        if self._result is not None:
            self._update_extrinsic_image_combo()

    @Slot(str)
    def _on_corner_error(self, error_msg: str):
        """角點偵測錯誤"""
        self.progress_bar.setVisible(False)
        self._set_buttons_enabled(True)
        QMessageBox.critical(self, "錯誤", f"角點偵測失敗：{error_msg}")
        self._corner_worker = None

    @Slot()
    def _on_calibrate(self):
        """執行標定（背景執行緒）"""
        if self.image_list.count() == 0:
            QMessageBox.warning(self, "提示", "請先載入圖像")
            return

        from vision_calib.core.intrinsic import IntrinsicCalibrationConfig
        from vision_calib.core.types import CheckerboardConfig
        from vision_calib.utils.worker import CalibrationWorker

        config = IntrinsicCalibrationConfig(
            checkerboard=CheckerboardConfig(
                rows=self.rows_spin.value(),
                cols=self.cols_spin.value(),
                square_size_mm=self.square_size_spin.value() * 10,  # cm → mm
            )
        )

        # 取得圖像路徑
        paths = []
        for i in range(self.image_list.count()):
            item = self.image_list.item(i)
            paths.append(item.data(Qt.UserRole))

        # 禁用按鈕
        self._set_buttons_enabled(False)

        # 顯示進度條
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.statusbar.showMessage("正在計算標定參數...")

        # 建立並啟動工作執行緒
        self._calib_worker = CalibrationWorker(paths, config, self)
        self._calib_worker.progress.connect(self._on_calib_progress)
        self._calib_worker.finished.connect(self._on_calib_finished)
        self._calib_worker.error.connect(self._on_calib_error)
        self._calib_worker.start()

    @Slot(int, int, str)
    def _on_calib_progress(self, current: int, total: int, message: str):
        """標定進度更新"""
        pct = int(current / total * 100) if total > 0 else 0
        self.progress_bar.setValue(pct)
        self.statusbar.showMessage(message)

    @Slot(object)
    def _on_calib_finished(self, result):
        """標定完成"""
        self.progress_bar.setVisible(False)
        self._set_buttons_enabled(True)
        self.statusbar.showMessage("就緒")
        self._calib_worker = None

        self._result = result
        self._display_calibration_result(result)
        self.export_btn.setEnabled(True)

        # 更新外參標定的圖像選擇下拉選單
        self._update_extrinsic_image_combo()

        QMessageBox.information(
            self,
            "標定完成",
            f"重投影誤差：{result.intrinsic.reprojection_error:.4f} 像素\n\n"
            f"相機矩陣和畸變係數已計算完成。\n"
            f"請點擊「匯出結果」儲存標定資料。",
        )

    @Slot(str)
    def _on_calib_error(self, error_msg: str):
        """標定錯誤"""
        self.progress_bar.setVisible(False)
        self._set_buttons_enabled(True)
        self.statusbar.showMessage("就緒")
        QMessageBox.critical(self, "錯誤", f"標定失敗：{error_msg}")
        logger.error(f"標定失敗：{error_msg}")
        self._calib_worker = None

    def _set_buttons_enabled(self, enabled: bool):
        """啟用/禁用操作按鈕"""
        self.detect_btn.setEnabled(enabled)
        self.calibrate_btn.setEnabled(enabled)
        self.add_images_btn.setEnabled(enabled)
        self.clear_images_btn.setEnabled(enabled)
        if enabled and self._result is not None:
            self.export_btn.setEnabled(True)
        elif not enabled:
            self.export_btn.setEnabled(False)

    def _display_calibration_result(self, result: CalibrationResult):
        """顯示標定結果"""
        intrinsic = result.intrinsic

        text = f"""標定完成！

══════════════════════════════════════
　相機內參矩陣 (Camera Matrix K)
══════════════════════════════════════

　　┌　{intrinsic.camera_matrix[0,0]:12.4f}　{intrinsic.camera_matrix[0,1]:12.4f}　{intrinsic.camera_matrix[0,2]:12.4f}　┐
　　│　{intrinsic.camera_matrix[1,0]:12.4f}　{intrinsic.camera_matrix[1,1]:12.4f}　{intrinsic.camera_matrix[1,2]:12.4f}　│
　　└　{intrinsic.camera_matrix[2,0]:12.4f}　{intrinsic.camera_matrix[2,1]:12.4f}　{intrinsic.camera_matrix[2,2]:12.4f}　┘

══════════════════════════════════════
　相機參數
══════════════════════════════════════

　　焦距 (fx)：{intrinsic.fx:.2f} pixels
　　焦距 (fy)：{intrinsic.fy:.2f} pixels
　　主點 (cx)：{intrinsic.cx:.2f} pixels
　　主點 (cy)：{intrinsic.cy:.2f} pixels

══════════════════════════════════════
　畸變係數 (Distortion Coefficients)
══════════════════════════════════════

　　k1 = {intrinsic.distortion_coeffs[0]:+.6f}
　　k2 = {intrinsic.distortion_coeffs[1]:+.6f}
　　p1 = {intrinsic.distortion_coeffs[2]:+.6f}
　　p2 = {intrinsic.distortion_coeffs[3]:+.6f}
　　k3 = {intrinsic.distortion_coeffs[4]:+.6f}

══════════════════════════════════════
　標定品質
══════════════════════════════════════

　　重投影誤差：{intrinsic.reprojection_error:.4f} pixels
　　圖像尺寸：{intrinsic.image_size[0]} × {intrinsic.image_size[1]}
　　使用圖像數：{result.num_images_used}
"""
        self.intrinsic_view.setText(text)

    @Slot()
    def _on_export(self):
        """匯出標定結果"""
        if self._result is None:
            QMessageBox.warning(self, "提示", "尚無標定結果可匯出")
            return

        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "匯出標定結果",
            "calibration",
            "HDF5 檔案 (*.h5);;MAT 檔案 (*.mat);;JSON 檔案 (*.json);;所有檔案 (*)",
        )

        if file_path:
            try:
                CalibrationFile.save(file_path, self._result)
                self.statusbar.showMessage(f"已匯出至：{file_path}")
                QMessageBox.information(
                    self,
                    "匯出成功",
                    f"標定結果已儲存至：\n{file_path}",
                )
            except Exception as e:
                QMessageBox.critical(self, "錯誤", f"匯出失敗：{e}")

    @Slot()
    def _on_open(self):
        """開啟標定檔案"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "開啟標定檔案",
            "",
            "標定檔案 (*.h5 *.mat *.json);;所有檔案 (*)",
        )

        if file_path:
            try:
                self._result = CalibrationFile.load(file_path)
                self._display_calibration_result(self._result)
                self.export_btn.setEnabled(True)
                self.statusbar.showMessage(f"已載入內參：{file_path}")
                logger.info(f"成功載入標定檔案，內參已就緒")

                # 更新外參頁面狀態提示
                self.ext_image_combo.clear()
                self.ext_image_combo.addItem("-- 請新增圖像並偵測角點 --")
                # 如果已有點位數據，保持按鈕啟用；否則禁用
                if len(self._point_data) >= 3:
                    self.ext_calibrate_btn.setEnabled(True)
                else:
                    self.ext_calibrate_btn.setEnabled(False)

                # 顯示提示訊息
                QMessageBox.information(
                    self,
                    "載入成功",
                    "內參已載入！\n\n"
                    "若要計算外參，請：\n"
                    "① 新增定位圖像\n"
                    "② 偵測角點\n"
                    "③ 切換到「外參計算」頁籤選擇圖像",
                )
            except Exception as e:
                QMessageBox.critical(self, "錯誤", f"無法載入檔案：{e}")
                logger.error(f"載入標定檔案失敗：{e}")

    @Slot()
    def _on_about(self):
        """顯示關於對話框"""
        QMessageBox.about(
            self,
            "關於 TSIC/CR-ICS01",
            f"""<h2>TSIC/CR-ICS01</h2>
            <h3>相機與手臂整合控制教學軟體</h3>
            <p>版本 {__version__}</p>
            <hr>
            <p>專業相機標定與座標轉換工具</p>
            <p>使用棋盤格圖案進行相機內外參標定，<br>
            支援像素、相機、世界座標系之間的轉換。</p>
            <hr>
            <p><b>開發單位：</b>TSIC</p>
            """,
        )

    def resizeEvent(self, event):
        """視窗縮放時調整圖像顯示"""
        super().resizeEvent(event)
        # 重新調整圖像檢視器的縮放
        if hasattr(self, 'image_scene') and self.image_scene.items():
            self.image_view.fitInView(
                self.image_scene.sceneRect(),
                Qt.KeepAspectRatio
            )

    def _update_extrinsic_image_combo(self):
        """更新外參標定的圖像選擇下拉選單"""
        self.ext_image_combo.clear()

        # 列出所有已載入的圖像（優先顯示已偵測角點的）
        all_images = []
        for i in range(self.image_list.count()):
            item = self.image_list.item(i)
            image_path = item.data(Qt.UserRole)
            has_corners = image_path in self._corner_cache and self._corner_cache[image_path] is not None
            all_images.append((Path(image_path).name, image_path, has_corners))

        if not all_images:
            self.ext_image_combo.addItem("-- 請先新增圖像 --")
            self.ext_calibrate_btn.setEnabled(False)
            return

        # 先加已偵測角點的，再加未偵測的
        for name, path, has_corners in all_images:
            if has_corners:
                self.ext_image_combo.addItem(f"✓ {name}", path)
            else:
                self.ext_image_combo.addItem(f"○ {name} (需偵測)", path)

        self.ext_calibrate_btn.setEnabled(True)

        detected_count = sum(1 for _, _, has in all_images if has)
        self.statusbar.showMessage(f"共 {len(all_images)} 張圖像，{detected_count} 張已偵測角點")

    def _on_algo_changed(self, index: int):
        """當 PnP 算法選擇變更時更新說明"""
        algo_descriptions = {
            "SOLVEPNP_ITERATIVE": "迭代法：最通用，適合大多數情況，需≥4個點",
            "SOLVEPNP_EPNP": "EPnP：高效率算法，點數較多(>10)時推薦，需≥4個點",
            "SOLVEPNP_P3P": "P3P：只需剛好3個點，可能有多解，適合點數極少的情況",
            "SOLVEPNP_AP3P": "AP3P：P3P改進版，數值穩定性更好，需≥3個點",
            "SOLVEPNP_IPPE": "IPPE：專為平面物體設計，適合棋盤格等平面標定，需≥4個點",
            "SOLVEPNP_IPPE_SQUARE": "IPPE_SQUARE：IPPE改進版，專為正方形標定板優化，需≥4個點",
        }
        algo_data = self.ext_algo_combo.currentData()
        desc = algo_descriptions.get(algo_data, "")
        if hasattr(self, 'algo_desc_label'):
            self.algo_desc_label.setText(desc)

        # 更新計算按鈕狀態
        self._check_extrinsic_ready()

    @Slot()
    def _on_calibrate_extrinsic(self):
        """執行外參標定"""
        if self._result is None:
            QMessageBox.warning(
                self,
                "尚無內參數據",
                "請先載入內參標定檔案（.h5/.mat/.json）\n"
                "或執行內參標定。"
            )
            return

        import cv2
        import numpy as np
        from vision_calib.core.types import CameraExtrinsic
        from vision_calib.core.extrinsic import ExtrinsicCalibrationResult
        from vision_calib.core.transform import CoordinateTransformer

        # 取得 PnP 算法
        algo_data = self.ext_algo_combo.currentData()
        algo_flags = {
            "SOLVEPNP_ITERATIVE": cv2.SOLVEPNP_ITERATIVE,
            "SOLVEPNP_EPNP": cv2.SOLVEPNP_EPNP,
            "SOLVEPNP_P3P": cv2.SOLVEPNP_P3P,
            "SOLVEPNP_AP3P": cv2.SOLVEPNP_AP3P,
            "SOLVEPNP_IPPE": cv2.SOLVEPNP_IPPE,
            "SOLVEPNP_IPPE_SQUARE": cv2.SOLVEPNP_IPPE_SQUARE,
        }
        pnp_flag = algo_flags.get(algo_data, cv2.SOLVEPNP_ITERATIVE)

        # 確定最少需要的點數
        min_points = 3 if algo_data in ("SOLVEPNP_P3P", "SOLVEPNP_AP3P") else 4

        try:
            # 根據選擇的方式進行計算
            if self.ext_use_points_radio.isChecked():
                # 方式一：使用點位數據
                # 優先使用 WebCAM 標記點（如果有有效的點）
                valid_webcam_points = [
                    p for p in self._marked_points
                    if p['world_x'] is not None and p['world_y'] is not None
                ]

                if len(valid_webcam_points) >= min_points:
                    # 使用 WebCAM 標記點
                    self.statusbar.showMessage("正在使用 WebCAM 標記點計算外參...")

                    object_points = np.array(
                        [[p['world_x'], p['world_y'], 0.0] for p in valid_webcam_points],
                        dtype=np.float32
                    )
                    image_points = np.array(
                        [[p['pixel_x'], p['pixel_y']] for p in valid_webcam_points],
                        dtype=np.float32
                    )
                    source = "webcam_points"
                    num_points = len(valid_webcam_points)
                elif len(self._point_data) >= min_points:
                    # 使用點位數據頁籤的數據
                    self.statusbar.showMessage("正在使用點位數據計算外參...")

                    object_points = np.array(
                        [[p[3], p[4], 0.0] for p in self._point_data],
                        dtype=np.float32
                    )
                    image_points = np.array(
                        [[p[1], p[2]] for p in self._point_data],
                        dtype=np.float32
                    )
                    source = "point_data"
                    num_points = len(self._point_data)
                else:
                    QMessageBox.warning(
                        self,
                        "點位不足",
                        f"選用的算法至少需要 {min_points} 個點位！\n"
                        f"WebCAM 標記點：{len(valid_webcam_points)} 個有效點\n"
                        f"點位數據頁籤：{len(self._point_data)} 個點\n\n"
                        "請在 WebCAM 頁籤標記點位並填入世界座標，\n"
                        "或前往「點位數據」頁籤添加點位。"
                    )
                    return

                # 執行 solvePnP
                success, rvec, tvec = cv2.solvePnP(
                    object_points,
                    image_points,
                    self._result.intrinsic.camera_matrix,
                    self._result.intrinsic.distortion_coeffs,
                    flags=pnp_flag,
                )

                if not success:
                    QMessageBox.critical(self, "失敗", "solvePnP 計算失敗！請檢查點位數據。")
                    return

                # 計算重投影誤差
                projected, _ = cv2.projectPoints(
                    object_points,
                    rvec,
                    tvec,
                    self._result.intrinsic.camera_matrix,
                    self._result.intrinsic.distortion_coeffs,
                )
                projected = projected.reshape(-1, 2)
                error = np.sqrt(np.mean(np.sum((image_points - projected) ** 2, axis=1)))

                # 建立結果
                extrinsic = CameraExtrinsic(
                    rotation_vector=rvec,
                    translation_vector=tvec,
                )
                self._extrinsic_result = ExtrinsicCalibrationResult(
                    extrinsic=extrinsic,
                    reprojection_error=error,
                    image_path=source,
                    num_points=num_points,
                )

            else:
                # 方式二：使用圖像角點
                image_path = self.ext_image_combo.currentData()
                if not image_path:
                    QMessageBox.warning(
                        self,
                        "無可用圖像",
                        "請先新增定位圖像，然後從下拉選單選擇。"
                    )
                    return

                # 取得角點
                corners = self._corner_cache.get(image_path)
                if corners is None:
                    # 自動偵測角點
                    self.statusbar.showMessage("正在偵測角點...")
                    from vision_calib.core.corner_detector import CornerDetector
                    from vision_calib.core.types import CheckerboardConfig

                    config = CheckerboardConfig(
                        rows=self.rows_spin.value(),
                        cols=self.cols_spin.value(),
                        square_size_mm=self.square_size_spin.value() * 10,
                    )
                    detector = CornerDetector(config)
                    result = detector.detect(image_path)

                    if not result.success:
                        QMessageBox.warning(
                            self,
                            "角點偵測失敗",
                            f"無法偵測到 {self.cols_spin.value()}×{self.rows_spin.value()} 棋盤格角點。"
                        )
                        return

                    corners = result.corners
                    self._corner_cache[image_path] = corners
                    self._update_extrinsic_image_combo()

                self.statusbar.showMessage("正在使用圖像角點計算外參...")

                from vision_calib.core.extrinsic import ExtrinsicCalibrator
                from vision_calib.core.types import CheckerboardConfig

                checkerboard = CheckerboardConfig(
                    rows=self.rows_spin.value(),
                    cols=self.cols_spin.value(),
                    square_size_mm=self.square_size_spin.value() * 10,
                )

                calibrator = ExtrinsicCalibrator(
                    intrinsic=self._result.intrinsic,
                    checkerboard=checkerboard,
                )
                self._extrinsic_result = calibrator.calibrate(image_path, corners)

            # 建立座標轉換器
            self._transformer = CoordinateTransformer(
                intrinsic=self._result.intrinsic,
                extrinsic=self._extrinsic_result.extrinsic,
            )

            # 顯示結果
            self._display_extrinsic_result()

            # 啟用按鈕
            self._enable_transform_buttons()
            self.ext_export_btn.setEnabled(True)

            self.statusbar.showMessage("外參計算完成")
            QMessageBox.information(
                self,
                "外參計算完成",
                f"重投影誤差：{self._extrinsic_result.reprojection_error:.4f} 像素\n\n"
                f"使用點位數：{self._extrinsic_result.num_points}\n"
                f"現在可以使用座標轉換功能。",
            )

        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"外參計算失敗：{e}")
            logger.error(f"外參計算失敗：{e}")

    def _display_extrinsic_result(self):
        """顯示外參標定結果"""
        if self._extrinsic_result is None:
            return

        result = self._extrinsic_result
        self.extrinsic_view.setText(result.summary())

    def _enable_transform_buttons(self):
        """啟用座標轉換按鈕"""
        self.p2w_btn.setEnabled(True)
        self.w2p_btn.setEnabled(True)
        self.transform_status.setText("座標轉換功能已就緒")

    @Slot()
    def _on_pixel_to_world(self):
        """像素座標 → 世界座標"""
        if self._transformer is None:
            QMessageBox.warning(self, "提示", "請先完成外參標定")
            return

        import numpy as np

        try:
            # 取得輸入
            u = self.pixel_u_spin.value()
            v = self.pixel_v_spin.value()
            z_world = self.world_z_spin.value()

            # 轉換
            pixel = np.array([u, v])
            world = self._transformer.pixel_to_world(pixel, z_world)

            # 顯示結果
            x, y, z = world[0], world[1], world[2]
            self.p2w_result.setText(f"X: {x:.2f} mm, Y: {y:.2f} mm, Z: {z:.2f} mm")
            self.statusbar.showMessage(f"像素 ({u:.1f}, {v:.1f}) → 世界 ({x:.2f}, {y:.2f}, {z:.2f})")

        except Exception as e:
            self.p2w_result.setText(f"轉換失敗：{e}")
            logger.error(f"像素→世界轉換失敗：{e}")

    @Slot()
    def _on_world_to_pixel(self):
        """世界座標 → 像素座標"""
        if self._transformer is None:
            QMessageBox.warning(self, "提示", "請先完成外參標定")
            return

        import numpy as np

        try:
            # 取得輸入
            x = self.world_x_spin.value()
            y = self.world_y_spin.value()
            z = self.world_z_input_spin.value()

            # 轉換
            world = np.array([x, y, z])
            pixel = self._transformer.world_to_pixel(world)

            # 顯示結果
            u, v = pixel[0], pixel[1]
            self.w2p_result.setText(f"U: {u:.1f} px, V: {v:.1f} px")
            self.statusbar.showMessage(f"世界 ({x:.2f}, {y:.2f}, {z:.2f}) → 像素 ({u:.1f}, {v:.1f})")

        except Exception as e:
            self.w2p_result.setText(f"轉換失敗：{e}")
            logger.error(f"世界→像素轉換失敗：{e}")

    # ===== 點位數據管理 =====

    @Slot()
    def _on_add_point(self):
        """添加單個點位"""
        point_id = self.point_id_spin.value()
        img_x = self.point_img_x_spin.value()
        img_y = self.point_img_y_spin.value()
        world_x = self.point_world_x_spin.value()
        world_y = self.point_world_y_spin.value()

        self._point_data.append([point_id, img_x, img_y, world_x, world_y])
        self._update_points_table()

        # 自動遞增 ID
        self.point_id_spin.setValue(point_id + 1)
        self.statusbar.showMessage(f"已添加點位 {point_id}")

    @Slot()
    def _on_clear_points(self):
        """清除所有點位數據"""
        if self._point_data:
            reply = QMessageBox.question(
                self, "確認清除",
                f"確定要清除所有 {len(self._point_data)} 個點位嗎？",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                self._point_data.clear()
                self._update_points_table()
                self.statusbar.showMessage("已清除所有點位")

    def _update_points_table(self):
        """更新點位數據表格"""
        from PySide6.QtWidgets import QTableWidgetItem

        self.points_table.setRowCount(len(self._point_data))
        for i, point in enumerate(self._point_data):
            for j, val in enumerate(point):
                item = QTableWidgetItem(f"{val:.2f}" if isinstance(val, float) else str(int(val)))
                self.points_table.setItem(i, j, item)

        # 更新標題
        self.points_list_group.setTitle(f"當前點位 (共 {len(self._point_data)} 個)")

        # 更新外參頁面狀態
        self.ext_points_status.setText(f"點位數據：{len(self._point_data)} 個點")

        # 如果有足夠點位且有內參數據，啟用計算外參按鈕
        if len(self._point_data) >= 3 and self._result is not None:
            self.ext_calibrate_btn.setEnabled(True)

    @Slot()
    def _on_import_points_csv(self):
        """匯入點位 CSV"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "匯入點位數據",
            "", "CSV 檔案 (*.csv);;所有檔案 (*)",
        )
        if not file_path:
            return

        try:
            import csv
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                count = 0
                for row in reader:
                    # 支持多種欄位名稱
                    point_id = float(row.get('id', row.get('ID', count + 1)))
                    img_x = float(row.get('image_x', row.get('img_x', row.get('x', 0))))
                    img_y = float(row.get('image_y', row.get('img_y', row.get('y', 0))))
                    world_x = float(row.get('world_x', row.get('X', 0)))
                    world_y = float(row.get('world_y', row.get('Y', 0)))
                    self._point_data.append([point_id, img_x, img_y, world_x, world_y])
                    count += 1

            self._update_points_table()
            self.statusbar.showMessage(f"已匯入 {count} 個點位")
            QMessageBox.information(self, "匯入成功", f"已匯入 {count} 個點位")
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"無法讀取 CSV：{e}")

    @Slot()
    def _on_export_points_csv(self):
        """匯出點位 CSV"""
        if not self._point_data:
            QMessageBox.warning(self, "無數據", "目前沒有點位數據可匯出")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "匯出點位數據",
            "point_data.csv", "CSV 檔案 (*.csv);;所有檔案 (*)",
        )
        if not file_path:
            return

        try:
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['id', 'image_x', 'image_y', 'world_x', 'world_y'])
                for point in self._point_data:
                    writer.writerow(point)

            self.statusbar.showMessage(f"已匯出至：{file_path}")
            QMessageBox.information(self, "匯出成功", f"已匯出 {len(self._point_data)} 個點位")
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", f"無法寫入 CSV：{e}")

    @Slot()
    def _on_export_excel_template(self):
        """匯出外參計算用的 Excel 範本"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "匯出Excel範本",
            "extrinsic_calibration_template.xlsx",
            "Excel 檔案 (*.xlsx);;所有檔案 (*)",
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "標定數據"

            # 定義樣式
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 表頭
            headers = [
                ("ID", "點位編號"),
                ("視覺X (像素)", "視覺辨識的X座標（像素）"),
                ("視覺Y (像素)", "視覺辨識的Y座標（像素）"),
                ("機械臂X (mm)", "機械臂座標系的X座標（毫米）"),
                ("機械臂Y (mm)", "機械臂座標系的Y座標（毫米）"),
            ]

            # 寫入表頭
            for col, (header, _) in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 設置欄寬
            col_widths = [8, 18, 18, 18, 18]
            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width

            # 添加範例數據行（可選，作為填寫示範）
            example_data = [
                [1, "", "", "", ""],
                [2, "", "", "", ""],
                [3, "", "", "", ""],
                [4, "", "", "", ""],
                [5, "", "", "", ""],
                [6, "", "", "", ""],
                [7, "", "", "", ""],
                [8, "", "", "", ""],
                [9, "", "", "", ""],
                [10, "", "", "", ""],
            ]

            data_alignment = Alignment(horizontal="center", vertical="center")
            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # 添加說明工作表
            ws_info = wb.create_sheet("使用說明")
            instructions = [
                ["外參計算 Excel 範本使用說明"],
                [""],
                ["1. 在「標定數據」工作表中填入對應點位的座標數據"],
                ["2. 視覺X/Y：在圖像中辨識到的點位像素座標"],
                ["3. 機械臂X/Y：機械臂移動到該點位時的座標（毫米）"],
                ["4. 至少需要 4 個點位才能計算外參"],
                ["5. 點位數量越多，計算結果越準確"],
                [""],
                ["注意事項："],
                ["- 確保視覺座標與機械臂座標一一對應"],
                ["- 點位應盡量分布在視野的不同區域"],
                ["- 避免所有點位在一條直線上"],
            ]
            for row_idx, row in enumerate(instructions, 1):
                cell = ws_info.cell(row=row_idx, column=1, value=row[0] if row else "")
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)

            ws_info.column_dimensions['A'].width = 60

            wb.save(file_path)
            self.statusbar.showMessage(f"已匯出Excel範本：{file_path}")
            QMessageBox.information(
                self, "匯出成功",
                f"已匯出Excel範本至：\n{file_path}\n\n"
                "請在「標定數據」工作表中填入座標數據後匯入。"
            )

        except ImportError:
            QMessageBox.critical(
                self, "缺少依賴",
                "需要安裝 openpyxl 套件：\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", f"無法建立Excel檔案：{e}")

    @Slot()
    def _on_import_excel_data(self):
        """匯入 Excel 數據用於外參計算"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "匯入Excel數據",
            "", "Excel 檔案 (*.xlsx *.xls);;所有檔案 (*)",
        )
        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)

            # 尋找數據工作表（優先「標定數據」，否則使用第一個）
            if "標定數據" in wb.sheetnames:
                ws = wb["標定數據"]
            else:
                ws = wb.active

            # 讀取數據（跳過表頭）
            imported_data = []
            valid_count = 0
            skip_count = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row is None or all(cell is None for cell in row):
                    continue

                try:
                    # 嘗試讀取各欄位
                    point_id = row[0] if row[0] is not None else row_idx - 1
                    img_x = row[1] if len(row) > 1 and row[1] is not None else None
                    img_y = row[2] if len(row) > 2 and row[2] is not None else None
                    world_x = row[3] if len(row) > 3 and row[3] is not None else None
                    world_y = row[4] if len(row) > 4 and row[4] is not None else None

                    # 檢查必要欄位是否有值
                    if img_x is None or img_y is None or world_x is None or world_y is None:
                        skip_count += 1
                        continue

                    # 轉換為數值
                    imported_data.append([
                        float(point_id),
                        float(img_x),
                        float(img_y),
                        float(world_x),
                        float(world_y)
                    ])
                    valid_count += 1

                except (ValueError, TypeError) as e:
                    skip_count += 1
                    continue

            wb.close()

            if not imported_data:
                QMessageBox.warning(
                    self, "無有效數據",
                    "Excel檔案中沒有找到有效的座標數據。\n\n"
                    "請確認：\n"
                    "1. 數據填寫在「標定數據」工作表\n"
                    "2. 視覺座標和機械臂座標都已填寫\n"
                    "3. 座標值為數字格式"
                )
                return

            # 更新點位數據
            self._point_data = imported_data
            self._update_points_table()

            # 更新外參計算頁籤的狀態
            self.excel_data_status.setText(f"已匯入 {valid_count} 個點位")
            self.ext_points_status.setText(f"點位數據：{valid_count} 個點")

            # 自動切換到使用點位數據模式
            self.ext_use_points_radio.setChecked(True)

            # 啟用計算外參按鈕
            self.ext_calibrate_btn.setEnabled(True)

            msg = f"已成功匯入 {valid_count} 個點位"
            if skip_count > 0:
                msg += f"\n（跳過 {skip_count} 個無效行）"

            self.statusbar.showMessage(msg)
            QMessageBox.information(self, "匯入成功", msg + "\n\n現在可以點擊「計算外參」進行計算。")

        except ImportError:
            QMessageBox.critical(
                self, "缺少依賴",
                "需要安裝 openpyxl 套件：\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"無法讀取Excel檔案：{e}")

    @Slot()
    def _on_load_corners_to_points(self):
        """從角點快取載入像素座標到點位數據"""
        # 找到第一個有角點的圖像
        corners_data = None
        for path, corners in self._corner_cache.items():
            if corners is not None:
                corners_data = corners
                break

        if corners_data is None:
            QMessageBox.warning(
                self, "無角點數據",
                "請先偵測圖像角點。\n\n"
                "① 載入圖像\n"
                "② 設定棋盤格參數\n"
                "③ 點擊「偵測角點」"
            )
            return

        # 如果有世界座標，合併；否則只載入像素座標
        if self._generated_world_coords:
            if len(corners_data) != len(self._generated_world_coords):
                QMessageBox.warning(
                    self, "數量不匹配",
                    f"角點數量 ({len(corners_data)}) 與世界座標數量 "
                    f"({len(self._generated_world_coords)}) 不匹配！\n\n"
                    "請確保棋盤格參數一致。"
                )
                return

            self._point_data.clear()
            for i, (corner, world) in enumerate(zip(corners_data, self._generated_world_coords)):
                point_id = world[0]
                img_x, img_y = corner[0], corner[1]
                world_x, world_y = world[1], world[2]
                self._point_data.append([point_id, img_x, img_y, world_x, world_y])
        else:
            # 只載入像素座標，世界座標設為 0
            self._point_data.clear()
            for i, corner in enumerate(corners_data):
                self._point_data.append([i + 1, corner[0], corner[1], 0.0, 0.0])

        self._update_points_table()
        self.statusbar.showMessage(f"已載入 {len(self._point_data)} 個點位")
        QMessageBox.information(
            self, "載入成功",
            f"已載入 {len(self._point_data)} 個點位到數據表。"
        )

    # ===== 匯出外參 =====

    @Slot()
    def _on_export_extrinsic(self):
        """匯出外參數據"""
        if self._extrinsic_result is None:
            QMessageBox.warning(self, "無數據", "請先完成外參計算")
            return

        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, "匯出外參",
            "extrinsic", "NPY 檔案 (*.npy);;JSON 檔案 (*.json);;所有檔案 (*)",
        )
        if not file_path:
            return

        try:
            import numpy as np
            import json

            ext = self._extrinsic_result.extrinsic
            rvec = ext.rotation_vector.flatten()
            tvec = ext.translation_vector.flatten()

            if file_path.endswith('.npy'):
                # 儲存為結構化陣列
                data = {
                    'rvec': ext.rotation_vector,
                    'tvec': ext.translation_vector,
                    'rotation_matrix': ext.rotation_matrix,
                }
                np.save(file_path, data)
            else:
                # 儲存為 JSON
                data = {
                    'rvec': rvec.tolist(),
                    'tvec': tvec.tolist(),
                    'rotation_matrix': ext.rotation_matrix.tolist(),
                    'reprojection_error': self._extrinsic_result.reprojection_error,
                }
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2)

            self.statusbar.showMessage(f"已匯出至：{file_path}")
            QMessageBox.information(self, "匯出成功", f"外參已儲存至：\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", f"無法儲存檔案：{e}")


def main():
    """應用程式入口點"""
    setup_logging()

    app = QApplication(sys.argv)
    app.setApplicationName("TSIC/CR-ICS01")
    app.setApplicationVersion(__version__)

    # 設置預設字型
    font = QFont()
    font.setFamily("Microsoft JhengHei UI")
    font.setPointSize(10)
    app.setFont(font)

    window = MainWindow()
    window.show()

    logger.info("應用程式啟動")
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
